#!/usr/bin/env python3
"""
Trinetra
Author: Sleeping Bhudda

Purpose
-------
Pull scan results from Nessus API, classify authentication status per host, show
an interactive GUI dashboard BEFORE exporting, then export Excel / CSV / PDF.

Core logic
----------
1. Total IPs are taken from the scan host inventory when available.
2. Raw auth plugin rows are grouped by Host + Auth Protocol.
3. Protocol status is calculated first.
4. Host status is calculated from protocol statuses.
5. Counts are mutually exclusive: PASS / FAIL / PARTIAL / NOCREDS / NOT_REACHABLE / UNKNOWN.

Supported evidence plugins
--------------------------
PASS:      141118, 110095, 122502, 117887, 19506 with Credentialed checks yes
FAIL:      104410, 122503, 91822
PARTIAL:   110385, 117885, 24786
NOCREDS:   110723
UNKNOWN:   117886, 21745, 110695 and no decisive evidence
NOT_REACHABLE: configured API scan target missing from result host inventory / CSV

Install
-------
Kali / Ubuntu:
    sudo apt update
    sudo apt install -y python3-tk
    python3 -m pip install requests openpyxl matplotlib

Run:
    python3 nessus_auth_rapid7_gui.py

Notes
-----
- Most Nessus scanners use self-signed TLS certificates. If your Nessus URL is
  https://<ip>:8834 and TLS verify fails, uncheck "Verify TLS".
- The script silently creates a temporary Nessus CSV export for preview. The
  GUI dashboard appears before final user export to Excel/PDF/CSV.
"""

from __future__ import annotations

import csv
import datetime as dt
import base64
import hashlib
import hmac
import ipaddress
import json
import math
import os
import secrets
import sys
import queue
import re
import subprocess
import tempfile
import threading
import time
import traceback
import warnings
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Optional, Tuple

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
except Exception as exc:  # pragma: no cover
    raise SystemExit("Tkinter is required. Install with: sudo apt install python3-tk") from exc

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False

try:
    import matplotlib
    matplotlib.use("TkAgg")
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    from matplotlib.backends.backend_pdf import PdfPages
    from matplotlib.figure import Figure
    MATPLOTLIB_AVAILABLE = True
except Exception:
    MATPLOTLIB_AVAILABLE = False

try:
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
except Exception:
    pass

warnings.filterwarnings("ignore", message="Unverified HTTPS request")

APP_NAME = "Trinetra"
APP_VERSION = "2.0"
AUTHOR = "Sleeping Bhudda"
USER_AGENT = f"nessus-auth-dashboard/{APP_VERSION} ({AUTHOR})"
AUTH_CONFIG_PATH = Path.home() / ".nessus_credential_assurance_auth.json"
AUTH_ITERATIONS = 260_000

# -----------------------------
# Classification constants
# -----------------------------

class AuthStatus(str, Enum):
    PASS = "PASS"
    FAIL = "FAIL"
    PARTIAL = "PARTIAL"
    NOCREDS = "NOCREDS"
    NOT_REACHABLE = "NOT_REACHABLE"
    UNKNOWN = "UNKNOWN"


PRIMARY_PASS = {141118, 110095, 122502}
PRIMARY_FAIL = {104410, 122503, 91822}
PRIMARY_PARTIAL = {110385, 117885, 24786}
PRIMARY_NOCREDS = {110723}
SECONDARY_PASS = {19506, 117887}
AMBIGUOUS_ONLY = {117886, 21745, 110695}
AUTH_RELATED_PLUGIN_IDS = (
    PRIMARY_PASS | PRIMARY_FAIL | PRIMARY_PARTIAL | PRIMARY_NOCREDS | SECONDARY_PASS | AMBIGUOUS_ONLY
)

PLUGIN_REASON = {
    19506: "Nessus Scan Information / Credentialed checks yes",
    141118: "Valid credentials provided for authentication protocol",
    110095: "Authentication success / no credential issues reported",
    122502: "Integration credential success",
    117887: "OS security patch assessment available / authenticated",
    104410: "Failure for provided credentials",
    122503: "Integration credential failure",
    91822: "Database authentication failure",
    110385: "Insufficient privilege after login",
    117885: "Intermittent authentication failure after login",
    24786: "Windows scan incomplete due to insufficient privileges",
    110723: "No credentials provided in scan policy",
    117886: "OS security patch assessment not available",
    21745: "OS security patch assessment failed / ambiguous login or other issue",
    110695: "Patch assessment checks not supported",
}

RECOMMENDATION_BY_STATUS = {
    AuthStatus.PASS: "No action required. Keep credential rotation and least privilege controls active.",
    AuthStatus.FAIL: "Validate username/password/key, domain format, network access, and protocol reachability. Re-test credentialed scan.",
    AuthStatus.PARTIAL: "Authentication is mixed or degraded. Check privilege level, sudo/admin rights, intermittent failures, and protocol-specific failures.",
    AuthStatus.NOCREDS: "Add valid credentials for the detected authentication protocol in the Nessus scan policy.",
    AuthStatus.NOT_REACHABLE: "Host was configured as a scan target but did not appear in the scan host inventory or CSV results. Verify routing, firewall, scanner reachability, and target scope.",
    AuthStatus.UNKNOWN: "Review raw plugin evidence. This may be unsupported OS/patch assessment, filtered output, or missing auth evidence.",
}

STATUS_ORDER = [AuthStatus.PASS, AuthStatus.FAIL, AuthStatus.PARTIAL, AuthStatus.NOCREDS, AuthStatus.NOT_REACHABLE, AuthStatus.UNKNOWN]
STATUS_LABELS = {
    AuthStatus.PASS: "Auth Passed",
    AuthStatus.FAIL: "Auth Failed",
    AuthStatus.PARTIAL: "Partial Auth",
    AuthStatus.NOCREDS: "No Credentials",
    AuthStatus.NOT_REACHABLE: "Not Reachable",
    AuthStatus.UNKNOWN: "Unknown",
}

STATUS_COLORS = {
    AuthStatus.PASS: "#22C55E",
    AuthStatus.FAIL: "#EF4444",
    AuthStatus.PARTIAL: "#F97316",
    AuthStatus.NOCREDS: "#8B5CF6",
    AuthStatus.NOT_REACHABLE: "#64748B",
    AuthStatus.UNKNOWN: "#EAB308",
}

METRIC_CARD_STYLES = {
    "Total IPs": ("TotalCard.TFrame", "TotalCardTitle.TLabel", "TotalCardValue.TLabel"),
    "Auth Passed": ("PassCard.TFrame", "PassCardTitle.TLabel", "PassCardValue.TLabel"),
    "Auth Failed": ("FailCard.TFrame", "FailCardTitle.TLabel", "FailCardValue.TLabel"),
    "Partial Auth": ("PartialCard.TFrame", "PartialCardTitle.TLabel", "PartialCardValue.TLabel"),
    "No Credentials": ("NoCredsCard.TFrame", "NoCredsCardTitle.TLabel", "NoCredsCardValue.TLabel"),
    "Not Reachable": ("NotReachableCard.TFrame", "NotReachableCardTitle.TLabel", "NotReachableCardValue.TLabel"),
    "Unknown": ("UnknownCard.TFrame", "UnknownCardTitle.TLabel", "UnknownCardValue.TLabel"),
    "Credential Coverage %": ("CoverageCard.TFrame", "CoverageCardTitle.TLabel", "CoverageCardValue.TLabel"),
    "Auth Success %": ("SuccessCard.TFrame", "SuccessCardTitle.TLabel", "SuccessCardValue.TLabel"),
}

PORT_TO_AUTH_PROTOCOL = {
    22: "SSH",
    2222: "SSH",
    445: "SMB",
    139: "SMB",
    135: "WMI",
    5985: "WINRM",
    5986: "WINRM",
    161: "SNMP",
    162: "SNMP",
    3389: "RDP",
    1433: "MSSQL",
    1521: "ORACLE",
    3306: "MYSQL",
    5432: "POSTGRESQL",
    27017: "MONGODB",
    6379: "REDIS",
}

PROTOCOL_KEYWORDS = [
    "SSH", "SMB", "WMI", "WINRM", "SNMP", "RDP", "MSSQL", "MYSQL", "POSTGRESQL",
    "POSTGRES", "ORACLE", "MONGODB", "REDIS", "DB", "DATABASE", "VSPHERE", "VCENTER", "ESX", "ESXI"
]

# -----------------------------
# Data model
# -----------------------------

@dataclass
class AuthFinding:
    host: str
    plugin_id: int
    plugin_name: str = ""
    plugin_output: str = ""
    port: Optional[int] = None
    transport_protocol: str = ""
    auth_protocol: str = "HOST"
    account: str = ""
    severity: str = ""
    risk: str = ""
    source: str = "csv"
    scan_id: str = ""
    scan_name: str = ""

    @property
    def reason(self) -> str:
        return PLUGIN_REASON.get(self.plugin_id, self.plugin_name or "Authentication-related plugin")

    @property
    def evidence_type(self) -> str:
        if self.plugin_id in PRIMARY_PASS or self.plugin_id in SECONDARY_PASS:
            return "PASS_EVIDENCE"
        if self.plugin_id in PRIMARY_FAIL:
            return "FAIL_EVIDENCE"
        if self.plugin_id in PRIMARY_PARTIAL:
            return "PARTIAL_EVIDENCE"
        if self.plugin_id in PRIMARY_NOCREDS:
            return "NOCREDS_EVIDENCE"
        return "AMBIGUOUS_EVIDENCE"


@dataclass
class ProtocolStatusRecord:
    host: str
    auth_protocol: str
    status: AuthStatus
    plugin_ids: List[int] = field(default_factory=list)
    reasons: List[str] = field(default_factory=list)
    accounts: List[str] = field(default_factory=list)
    ports: List[str] = field(default_factory=list)
    scan_ids: List[str] = field(default_factory=list)
    scan_names: List[str] = field(default_factory=list)
    confidence: str = "High"
    evidence_count: int = 0


@dataclass
class HostStatusRecord:
    host: str
    status: AuthStatus
    protocols: Dict[str, AuthStatus] = field(default_factory=dict)
    plugin_ids: List[int] = field(default_factory=list)
    reasons: List[str] = field(default_factory=list)
    accounts: List[str] = field(default_factory=list)
    scan_ids: List[str] = field(default_factory=list)
    scan_names: List[str] = field(default_factory=list)
    recommendation: str = ""
    pass_protocols: List[str] = field(default_factory=list)
    fail_protocols: List[str] = field(default_factory=list)
    partial_protocols: List[str] = field(default_factory=list)
    nocreds_protocols: List[str] = field(default_factory=list)
    unknown_protocols: List[str] = field(default_factory=list)


@dataclass
class DashboardData:
    scan_name: str = ""
    scan_id: str = ""
    history_id: str = ""
    generated_at: str = ""
    source_file: str = ""
    authoritative_hosts: List[str] = field(default_factory=list)
    findings: List[AuthFinding] = field(default_factory=list)
    protocol_records: List[ProtocolStatusRecord] = field(default_factory=list)
    host_records: List[HostStatusRecord] = field(default_factory=list)
    metrics: Dict[str, Any] = field(default_factory=dict)
    raw_rows_count: int = 0
    notes: List[str] = field(default_factory=list)

# -----------------------------
# Utility helpers
# -----------------------------

def now_string() -> str:
    return dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def format_nessus_time(value: Any) -> str:
    if value is None or value == "":
        return ""

    try:
        if isinstance(value, str):
            text = value.strip()
            if not text:
                return ""
            if re.fullmatch(r"\d+(\.\d+)?", text):
                value = float(text)
            else:
                iso_text = text.replace("Z", "+00:00")
                parsed = dt.datetime.fromisoformat(iso_text)
                if parsed.tzinfo:
                    parsed = parsed.astimezone()
                return parsed.strftime("%Y-%m-%d %H:%M:%S %Z").strip()

        if isinstance(value, (int, float)):
            timestamp = float(value)
            if timestamp > 10_000_000_000_000:
                timestamp /= 1_000_000
            elif timestamp > 10_000_000_000:
                timestamp /= 1_000
            return dt.datetime.fromtimestamp(timestamp).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return str(value)

    return str(value)


def safe_int(value: Any) -> Optional[int]:
    try:
        if value is None:
            return None
        s = str(value).strip()
        if not s:
            return None
        return int(float(s))
    except Exception:
        return None


def unique_preserve_order(values: Iterable[Any]) -> List[Any]:
    seen = set()
    out = []
    for v in values:
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue
        if s not in seen:
            seen.add(s)
            out.append(s)
    return out


def normalize_header(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", (name or "").strip().lower()).strip("_")


def first_present(row: Dict[str, Any], *keys: str, default: str = "") -> str:
    normalized = {normalize_header(k): v for k, v in row.items()}
    for key in keys:
        nk = normalize_header(key)
        if nk in normalized and normalized[nk] is not None:
            return str(normalized[nk])
    return default


def sort_hosts(hosts: Iterable[str]) -> List[str]:
    def key_func(h: str):
        try:
            ip = ipaddress.ip_address(h)
            return (0, ip.version, int(ip))
        except Exception:
            return (1, 0, str(h))
    return sorted(set([str(h).strip() for h in hosts if str(h).strip()]), key=key_func)


def normalize_host_token(value: Any) -> str:
    return str(value or "").strip().strip(",;")


def split_target_text(value: Any) -> List[str]:
    text = str(value or "")
    parts = re.split(r"[\s,;]+", text)
    return [normalize_host_token(part) for part in parts if normalize_host_token(part)]


def expand_target_token(token: str, max_expand: int = 65536) -> List[str]:
    token = normalize_host_token(token)
    if not token:
        return []

    try:
        network = ipaddress.ip_network(token, strict=False)
        if network.num_addresses > max_expand:
            return [token]
        return [str(ip) for ip in network.hosts()] or [str(network.network_address)]
    except ValueError:
        pass

    if "-" in token:
        start, end = token.split("-", 1)
        start = start.strip()
        end = end.strip()
        try:
            start_ip = ipaddress.ip_address(start)
            if re.fullmatch(r"\d+", end):
                prefix = start.rsplit(".", 1)[0] if start_ip.version == 4 else ""
                if prefix:
                    end_ip = ipaddress.ip_address(f"{prefix}.{end}")
                else:
                    end_ip = ipaddress.ip_address(end)
            else:
                end_ip = ipaddress.ip_address(end)
            if start_ip.version == end_ip.version and int(start_ip) <= int(end_ip):
                count = int(end_ip) - int(start_ip) + 1
                if count <= max_expand:
                    return [str(ipaddress.ip_address(int(start_ip) + i)) for i in range(count)]
        except ValueError:
            pass

    return [token]


def extract_regex_line(patterns: List[str], text: str) -> str:
    for pattern in patterns:
        m = re.search(pattern, text or "", flags=re.IGNORECASE | re.MULTILINE)
        if m:
            return m.group(1).strip()
    return ""


def infer_auth_protocol(plugin_id: int, plugin_name: str, plugin_output: str, port: Optional[int], transport_protocol: str) -> str:
    output = plugin_output or ""
    name = plugin_name or ""

    # Prefer explicit protocol/account fields from plugin output.
    proto = extract_regex_line([
        r"^\s*authentication\s+protocol\s*[:=]\s*([A-Za-z0-9_ ./\\-]+)\s*$",
        r"^\s*auth\s+protocol\s*[:=]\s*([A-Za-z0-9_ ./\\-]+)\s*$",
        r"^\s*protocol\s*[:=]\s*([A-Za-z0-9_ ./\\-]+)\s*$",
        r"^\s*service\s*[:=]\s*([A-Za-z0-9_ ./\\-]+)\s*$",
    ], output)

    if proto:
        proto_clean = proto.upper().replace(" ", "_")
        if proto_clean in {"TCP", "UDP", "ICMP"}:
            proto = ""
        else:
            return normalize_auth_protocol(proto_clean)

    combined = f"{name}\n{output}".upper()
    for kw in PROTOCOL_KEYWORDS:
        if re.search(r"\b" + re.escape(kw) + r"\b", combined):
            return normalize_auth_protocol(kw)

    if port in PORT_TO_AUTH_PROTOCOL:
        return PORT_TO_AUTH_PROTOCOL[port]

    # 19506 is a host-level scan-info plugin.
    if plugin_id == 19506:
        return "HOST"

    # Patch/integration plugins may not expose protocol in CSV.
    if plugin_id in {122502, 122503}:
        return "INTEGRATION"

    if plugin_id == 91822:
        return "DATABASE"

    return "HOST"


def normalize_auth_protocol(proto: str) -> str:
    p = (proto or "").strip().upper().replace(" ", "_").replace("-", "_")
    aliases = {
        "WINDOWS": "SMB/WMI",
        "MICROSOFT_WINDOWS": "SMB/WMI",
        "CIFS": "SMB",
        "WINRM_HTTP": "WINRM",
        "WINRM_HTTPS": "WINRM",
        "POSTGRES": "POSTGRESQL",
        "DB": "DATABASE",
        "DATABASE": "DATABASE",
        "VSPHERE": "VCENTER/ESX",
        "VCENTER": "VCENTER/ESX",
        "ESX": "VCENTER/ESX",
        "ESXI": "VCENTER/ESX",
    }
    return aliases.get(p, p or "HOST")


def extract_account(plugin_output: str) -> str:
    return extract_regex_line([
        r"^\s*account\s*[:=]\s*(.+?)\s*$",
        r"^\s*username\s*[:=]\s*(.+?)\s*$",
        r"^\s*user\s*[:=]\s*(.+?)\s*$",
        r"^\s*login\s*[:=]\s*(.+?)\s*$",
    ], plugin_output or "")


def credentialed_checks_yes(text: str) -> bool:
    if not text:
        return False
    normalized = re.sub(r"\s+", " ", text.strip().lower())
    patterns = [
        r"credentialed\s+checks\s*:\s*yes",
        r"credentialed\s+checks\s*=\s*yes",
        r"credentialed\s+checks\s+yes",
    ]
    return any(re.search(p, normalized, flags=re.IGNORECASE) for p in patterns)

# -----------------------------
# Nessus API client
# -----------------------------

class NessusClient:
    def __init__(self, base_url: str, access_key: str, secret_key: str, verify_tls: bool = False):
        self.base_url = base_url.rstrip("/")
        self.session = requests.Session()
        self.session.headers.update({
            "X-ApiKeys": f"accessKey={access_key}; secretKey={secret_key}",
            "Accept": "application/json",
            "Content-Type": "application/json",
            "User-Agent": USER_AGENT,
        })
        self.verify_tls = verify_tls
        retry = Retry(
            total=3,
            backoff_factor=1,
            status_forcelist=[429, 500, 502, 503, 504],
            allowed_methods=["GET", "POST"],
        )
        adapter = HTTPAdapter(max_retries=retry)
        self.session.mount("https://", adapter)
        self.session.mount("http://", adapter)

    def request(self, method: str, path: str, **kwargs) -> requests.Response:
        url = f"{self.base_url}{path}"
        kwargs.setdefault("verify", self.verify_tls)
        kwargs.setdefault("timeout", 120)
        r = self.session.request(method, url, **kwargs)
        if r.status_code >= 400:
            detail = r.text[:1000]
            raise RuntimeError(f"Nessus API error {r.status_code} for {method} {path}: {detail}")
        return r

    def list_scan_inventory(self) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        data = self.request("GET", "/scans").json()
        scans = data.get("scans", []) or []
        folders = data.get("folders", []) or []
        folder_names = {str(f.get("id")): f.get("name", "") for f in folders}
        for s in scans:
            fid = str(s.get("folder_id", ""))
            s["folder_name"] = folder_names.get(fid, "")
        return scans, folders

    def list_scans(self) -> List[Dict[str, Any]]:
        scans, _folders = self.list_scan_inventory()
        return scans

    def get_scan_details(self, scan_id: str, history_id: Optional[str] = None) -> Dict[str, Any]:
        params = {}
        if history_id:
            params["history_id"] = history_id
        return self.request("GET", f"/scans/{scan_id}", params=params).json()

    def export_scan_csv(self, scan_id: str, dest_path: Path, history_id: Optional[str] = None,
                        progress_cb: Optional[Callable[[str], None]] = None) -> Path:
        params = {}
        if history_id:
            params["history_id"] = history_id
        body = {"format": "csv"}
        if progress_cb:
            progress_cb("Requesting temporary Nessus CSV export for preview...")
        data = self.request("POST", f"/scans/{scan_id}/export", params=params, json=body).json()
        file_id = data.get("file") or data.get("file_id")
        if file_id is None:
            raise RuntimeError(f"Export request did not return file id: {data}")

        for i in range(180):
            status_data = self.request("GET", f"/scans/{scan_id}/export/{file_id}/status").json()
            status = (status_data.get("status") or "").lower()
            if progress_cb and i % 3 == 0:
                progress_cb(f"Nessus export status: {status or 'waiting'}")
            if status == "ready":
                break
            if status in {"error", "canceled", "cancelled"}:
                raise RuntimeError(f"Nessus export failed: {status_data}")
            time.sleep(2)
        else:
            raise TimeoutError("Nessus CSV export did not become ready in time.")

        if progress_cb:
            progress_cb("Downloading temporary CSV export...")
        # Download as binary; endpoint often returns octet-stream.
        r = self.request("GET", f"/scans/{scan_id}/export/{file_id}/download", headers={"Accept": "application/octet-stream"}, stream=True)
        with dest_path.open("wb") as f:
            for chunk in r.iter_content(chunk_size=1024 * 128):
                if chunk:
                    f.write(chunk)
        return dest_path

# -----------------------------
# Parser and classifier
# -----------------------------

class AuthClassifier:
    def __init__(self):
        self.notes: List[str] = []
        self.configured_targets: List[str] = []
        self.host_source_ids: Dict[str, List[str]] = defaultdict(list)
        self.host_source_names: Dict[str, List[str]] = defaultdict(list)
        self.target_source_ids: Dict[str, List[str]] = defaultdict(list)
        self.target_source_names: Dict[str, List[str]] = defaultdict(list)

    def remember_host_sources(self, hosts: Iterable[str], scan_id: str, scan_name: str) -> None:
        for host in hosts:
            host = str(host or "").strip()
            if not host:
                continue
            if scan_id and scan_id not in self.host_source_ids[host]:
                self.host_source_ids[host].append(scan_id)
            if scan_name and scan_name not in self.host_source_names[host]:
                self.host_source_names[host].append(scan_name)

    def remember_target_sources(self, hosts: Iterable[str], scan_id: str, scan_name: str) -> None:
        for host in hosts:
            host = str(host or "").strip()
            if not host:
                continue
            if scan_id and scan_id not in self.target_source_ids[host]:
                self.target_source_ids[host].append(scan_id)
            if scan_name and scan_name not in self.target_source_names[host]:
                self.target_source_names[host].append(scan_name)

    def parse_csv(self, csv_path: Path, scan_id: str = "", scan_name: str = "") -> Tuple[List[Dict[str, Any]], List[AuthFinding], List[str]]:
        raw_rows: List[Dict[str, Any]] = []
        findings: List[AuthFinding] = []
        all_hosts: List[str] = []

        # Try utf-8-sig, then fallback latin-1.
        text_encoding = "utf-8-sig"
        try:
            with csv_path.open("r", encoding=text_encoding, newline="") as f:
                sample = f.read(4096)
        except UnicodeDecodeError:
            text_encoding = "latin-1"

        with csv_path.open("r", encoding=text_encoding, newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                raw_rows.append(row)
                host = first_present(row, "Host", "Host IP", "IP", "Hostname", "DNS Name", "Asset IP")
                host = host.strip()
                if host:
                    all_hosts.append(host)

                plugin_id = safe_int(first_present(row, "Plugin ID", "PluginID", "Plugin"))
                if plugin_id is None or plugin_id not in AUTH_RELATED_PLUGIN_IDS:
                    continue

                plugin_output = first_present(row, "Plugin Output", "PluginOutput", "Output", "Synopsis", default="")
                plugin_name = first_present(row, "Name", "Plugin Name", "Plugin", default="")

                # Avoid using 19506 as evidence unless it explicitly contains Credentialed checks yes.
                if plugin_id == 19506 and not credentialed_checks_yes(plugin_output):
                    continue

                port = safe_int(first_present(row, "Port", "Service Port", default=""))
                transport_protocol = first_present(row, "Protocol", "Transport", default="")
                auth_protocol = infer_auth_protocol(plugin_id, plugin_name, plugin_output, port, transport_protocol)
                account = extract_account(plugin_output)
                severity = first_present(row, "Severity", "Risk", "Risk Factor", default="")
                risk = first_present(row, "Risk", "Risk Factor", default=severity)

                if host:
                    findings.append(AuthFinding(
                        host=host,
                        plugin_id=plugin_id,
                        plugin_name=plugin_name,
                        plugin_output=plugin_output,
                        port=port,
                        transport_protocol=transport_protocol,
                        auth_protocol=auth_protocol,
                        account=account,
                        severity=severity,
                        risk=risk,
                        source=str(csv_path),
                        scan_id=scan_id,
                        scan_name=scan_name,
                    ))

        return raw_rows, findings, sort_hosts(all_hosts)

    def hosts_from_scan_details(self, details: Dict[str, Any]) -> List[str]:
        hosts = []
        for item in details.get("hosts", []) or []:
            candidate = (
                item.get("hostname") or item.get("host-fqdn") or item.get("host_ip") or
                item.get("ip") or item.get("name") or item.get("dns_name") or ""
            )
            candidate = str(candidate).strip()
            if candidate:
                hosts.append(candidate)
        return sort_hosts(hosts)

    def targets_from_scan_details(self, details: Dict[str, Any]) -> List[str]:
        raw_targets: List[str] = []
        settings = details.get("settings", {}) or {}
        info = details.get("info", {}) or {}

        for container in (settings, info, details):
            for key in ("text_targets", "targets", "target", "scan_targets"):
                value = container.get(key) if isinstance(container, dict) else None
                if isinstance(value, list):
                    raw_targets.extend(str(v) for v in value)
                elif value:
                    raw_targets.extend(split_target_text(value))

        expanded: List[str] = []
        for target in raw_targets:
            expanded.extend(expand_target_token(target))
        return sort_hosts(expanded)

    def classify_protocol(self, findings: List[AuthFinding]) -> ProtocolStatusRecord:
        if not findings:
            return ProtocolStatusRecord(host="", auth_protocol="HOST", status=AuthStatus.UNKNOWN)

        host = findings[0].host
        auth_protocol = findings[0].auth_protocol or "HOST"
        ids = {f.plugin_id for f in findings}

        # Strict precedence: partial first because it means login was degraded/unstable.
        if ids & PRIMARY_PARTIAL:
            status = AuthStatus.PARTIAL
        elif (ids & PRIMARY_PASS) and (ids & (PRIMARY_FAIL | PRIMARY_NOCREDS)):
            status = AuthStatus.PARTIAL
        elif ids & PRIMARY_PASS:
            status = AuthStatus.PASS
        elif ids & PRIMARY_FAIL:
            status = AuthStatus.FAIL
        elif ids & PRIMARY_NOCREDS:
            status = AuthStatus.NOCREDS
        elif 19506 in ids and any(credentialed_checks_yes(f.plugin_output) for f in findings):
            status = AuthStatus.PASS
        elif 117887 in ids:
            status = AuthStatus.PASS
        else:
            status = AuthStatus.UNKNOWN

        confidence = "High"
        if status == AuthStatus.PASS and not (ids & PRIMARY_PASS):
            confidence = "Medium - fallback evidence"
        if status == AuthStatus.UNKNOWN:
            confidence = "Low - ambiguous or missing evidence"

        return ProtocolStatusRecord(
            host=host,
            auth_protocol=auth_protocol,
            status=status,
            plugin_ids=sorted(ids),
            reasons=unique_preserve_order([f.reason for f in findings]),
            accounts=unique_preserve_order([f.account for f in findings]),
            ports=unique_preserve_order([str(f.port) for f in findings if f.port is not None]),
            scan_ids=unique_preserve_order([f.scan_id for f in findings]),
            scan_names=unique_preserve_order([f.scan_name for f in findings]),
            confidence=confidence,
            evidence_count=len(findings),
        )

    def classify(self, scan_name: str, scan_id: str, history_id: str,
                 authoritative_hosts: List[str], findings: List[AuthFinding], raw_rows_count: int,
                 source_file: str = "") -> DashboardData:
        if not authoritative_hosts:
            authoritative_hosts = sort_hosts([f.host for f in findings])
            self.notes.append("Authoritative host inventory was not available; Total IPs falls back to unique hosts in CSV.")

        result_hosts = sort_hosts(set(authoritative_hosts) | {f.host for f in findings})
        configured_targets = getattr(self, "configured_targets", [])
        if configured_targets:
            self.notes.append(
                "Configured scan targets were read from Nessus scan settings. NOT_REACHABLE means the target did not appear in the host inventory or CSV results."
            )
            self.notes.append(
                "If targets are hostnames, DNS aliases, asset groups, or very large CIDR ranges, compare results carefully because Nessus may report a different resolved host value."
            )
            authoritative_hosts = sort_hosts(set(result_hosts) | set(configured_targets))

        by_host_protocol: Dict[str, Dict[str, List[AuthFinding]]] = defaultdict(lambda: defaultdict(list))
        for f in findings:
            proto = f.auth_protocol or "HOST"
            by_host_protocol[f.host][proto].append(f)

        protocol_records: List[ProtocolStatusRecord] = []
        protocol_records_by_host: Dict[str, List[ProtocolStatusRecord]] = defaultdict(list)
        for host, proto_map in by_host_protocol.items():
            for proto, f_list in proto_map.items():
                rec = self.classify_protocol(f_list)
                protocol_records.append(rec)
                protocol_records_by_host[host].append(rec)

        host_records: List[HostStatusRecord] = []
        for host in authoritative_hosts:
            protos = protocol_records_by_host.get(host, [])
            if configured_targets and host in configured_targets and host not in result_hosts:
                status = AuthStatus.NOT_REACHABLE
                proto_status_map = {}
                plugin_ids = []
                reasons = ["Configured scan target did not appear in Nessus host inventory or CSV results"]
                accounts = []
                scan_ids = unique_preserve_order(self.target_source_ids.get(host, []))
                scan_names = unique_preserve_order(self.target_source_names.get(host, []))
            elif not protos:
                status = AuthStatus.UNKNOWN
                proto_status_map = {}
                plugin_ids: List[int] = []
                reasons = ["No authentication evidence found in scan export"]
                accounts: List[str] = []
                scan_ids = unique_preserve_order(self.host_source_ids.get(host, []) + self.target_source_ids.get(host, []))
                scan_names = unique_preserve_order(self.host_source_names.get(host, []) + self.target_source_names.get(host, []))
            else:
                proto_status_map = {p.auth_protocol: p.status for p in protos}
                decisive_states = {p.status for p in protos if p.status != AuthStatus.UNKNOWN}
                plugin_ids = sorted(set(pid for p in protos for pid in p.plugin_ids))
                reasons = unique_preserve_order([r for p in protos for r in p.reasons])
                accounts = unique_preserve_order([a for p in protos for a in p.accounts])
                scan_ids = unique_preserve_order([sid for p in protos for sid in p.scan_ids])
                scan_names = unique_preserve_order([name for p in protos for name in p.scan_names])

                if not decisive_states:
                    status = AuthStatus.UNKNOWN
                elif AuthStatus.PARTIAL in decisive_states:
                    status = AuthStatus.PARTIAL
                elif len(decisive_states) > 1:
                    status = AuthStatus.PARTIAL
                elif decisive_states == {AuthStatus.PASS}:
                    status = AuthStatus.PASS
                elif decisive_states == {AuthStatus.FAIL}:
                    status = AuthStatus.FAIL
                elif decisive_states == {AuthStatus.NOCREDS}:
                    status = AuthStatus.NOCREDS
                else:
                    status = AuthStatus.UNKNOWN

            host_records.append(HostStatusRecord(
                host=host,
                status=status,
                protocols=proto_status_map,
                plugin_ids=plugin_ids,
                reasons=reasons,
                accounts=accounts,
                scan_ids=scan_ids,
                scan_names=scan_names,
                recommendation=RECOMMENDATION_BY_STATUS.get(status, "Review evidence."),
                pass_protocols=sorted([p.auth_protocol for p in protos if p.status == AuthStatus.PASS]),
                fail_protocols=sorted([p.auth_protocol for p in protos if p.status == AuthStatus.FAIL]),
                partial_protocols=sorted([p.auth_protocol for p in protos if p.status == AuthStatus.PARTIAL]),
                nocreds_protocols=sorted([p.auth_protocol for p in protos if p.status == AuthStatus.NOCREDS]),
                unknown_protocols=sorted([p.auth_protocol for p in protos if p.status == AuthStatus.UNKNOWN]),
            ))

        host_records.sort(key=lambda h: sort_hosts([h.host])[0] if h.host else "")

        counts = Counter([h.status for h in host_records])
        total = len(authoritative_hosts)
        passed = counts.get(AuthStatus.PASS, 0)
        failed = counts.get(AuthStatus.FAIL, 0)
        partial = counts.get(AuthStatus.PARTIAL, 0)
        nocreds = counts.get(AuthStatus.NOCREDS, 0)
        not_reachable = counts.get(AuthStatus.NOT_REACHABLE, 0)
        unknown = counts.get(AuthStatus.UNKNOWN, 0)

        metrics = {
            "Total IPs": total,
            "Auth Passed": passed,
            "Auth Failed": failed,
            "Partial Auth": partial,
            "No Credentials": nocreds,
            "Not Reachable": not_reachable,
            "Unknown": unknown,
            "Auth Success %": round((passed / total * 100), 2) if total else 0.0,
            "Auth Failure %": round((failed / total * 100), 2) if total else 0.0,
            "Credential Coverage %": round(((passed + partial) / total * 100), 2) if total else 0.0,
            "Total Auth Findings": len(findings),
            "Raw Rows Parsed": raw_rows_count,
        }

        data = DashboardData(
            scan_name=scan_name,
            scan_id=str(scan_id or ""),
            history_id=str(history_id or ""),
            generated_at=now_string(),
            source_file=source_file,
            authoritative_hosts=authoritative_hosts,
            findings=findings,
            protocol_records=protocol_records,
            host_records=host_records,
            metrics=metrics,
            raw_rows_count=raw_rows_count,
            notes=list(self.notes),
        )
        return data

# -----------------------------
# Exporters
# -----------------------------

class Exporter:
    @staticmethod
    def host_rows(data: DashboardData) -> List[Dict[str, Any]]:
        rows = []
        for h in data.host_records:
            rows.append({
                "Host": h.host,
                "Source Scan IDs": ", ".join(h.scan_ids),
                "Source Scan Names": ", ".join(h.scan_names),
                "Status": h.status.value,
                "Pass Protocols": ", ".join(h.pass_protocols),
                "Fail Protocols": ", ".join(h.fail_protocols),
                "Partial Protocols": ", ".join(h.partial_protocols),
                "NoCreds Protocols": ", ".join(h.nocreds_protocols),
                "Unknown Protocols": ", ".join(h.unknown_protocols),
                "Plugin IDs": ", ".join(map(str, h.plugin_ids)),
                "Reasons": " | ".join(h.reasons),
                "Accounts": ", ".join(h.accounts),
                "Recommendation": h.recommendation,
            })
        return rows

    @staticmethod
    def protocol_rows(data: DashboardData) -> List[Dict[str, Any]]:
        rows = []
        for p in data.protocol_records:
            rows.append({
                "Host": p.host,
                "Source Scan IDs": ", ".join(p.scan_ids),
                "Source Scan Names": ", ".join(p.scan_names),
                "Auth Protocol": p.auth_protocol,
                "Status": p.status.value,
                "Plugin IDs": ", ".join(map(str, p.plugin_ids)),
                "Reasons": " | ".join(p.reasons),
                "Accounts": ", ".join(p.accounts),
                "Ports": ", ".join(p.ports),
                "Confidence": p.confidence,
                "Evidence Count": p.evidence_count,
            })
        return rows

    @staticmethod
    def finding_rows(data: DashboardData) -> List[Dict[str, Any]]:
        rows = []
        for f in data.findings:
            rows.append({
                "Host": f.host,
                "Source Scan ID": f.scan_id,
                "Source Scan Name": f.scan_name,
                "Auth Protocol": f.auth_protocol,
                "Plugin ID": f.plugin_id,
                "Plugin Name": f.plugin_name,
                "Evidence Type": f.evidence_type,
                "Reason": f.reason,
                "Port": f.port if f.port is not None else "",
                "Transport Protocol": f.transport_protocol,
                "Account": f.account,
                "Plugin Output": f.plugin_output,
            })
        return rows

    @staticmethod
    def write_csv(path: Path, rows: List[Dict[str, Any]]) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        if not rows:
            path.write_text("No data\n", encoding="utf-8")
            return
        with path.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            writer.writerows(rows)

    @staticmethod
    def export_csv_bundle(data: DashboardData, folder: Path) -> Path:
        folder.mkdir(parents=True, exist_ok=True)
        Exporter.write_csv(folder / "host_status.csv", Exporter.host_rows(data))
        Exporter.write_csv(folder / "protocol_status.csv", Exporter.protocol_rows(data))
        Exporter.write_csv(folder / "auth_findings.csv", Exporter.finding_rows(data))
        for status in STATUS_ORDER:
            rows = [r for r in Exporter.host_rows(data) if r["Status"] == status.value]
            Exporter.write_csv(folder / f"{status.value.lower()}_hosts.csv", rows)
        summary = {
            "scan_name": data.scan_name,
            "scan_id": data.scan_id,
            "history_id": data.history_id,
            "generated_at": data.generated_at,
            "metrics": data.metrics,
            "notes": data.notes,
        }
        (folder / "summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")
        return folder

    @staticmethod
    def export_excel(data: DashboardData, path: Path) -> None:
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl not installed. Install: python3 -m pip install openpyxl")

        wb = Workbook()
        ws = wb.active
        ws.title = "Dashboard"

        fills = {
            "title": PatternFill("solid", fgColor="1F2937"),
            "header": PatternFill("solid", fgColor="D9EAF7"),
            "pass": PatternFill("solid", fgColor="C6EFCE"),
            "fail": PatternFill("solid", fgColor="FFC7CE"),
            "partial": PatternFill("solid", fgColor="FCE4D6"),
            "nocreds": PatternFill("solid", fgColor="E7E6E6"),
            "not_reachable": PatternFill("solid", fgColor="D9E2F3"),
            "unknown": PatternFill("solid", fgColor="FFF2CC"),
        }
        thin = Side(style="thin", color="B7B7B7")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.merge_cells("A1:H1")
        ws["A1"] = f"{APP_NAME} v{APP_VERSION}"
        ws["A1"].font = Font(color="FFFFFF", bold=True, size=16)
        ws["A1"].fill = fills["title"]
        ws["A1"].alignment = Alignment(horizontal="center")

        meta_rows = [
            ("Author", AUTHOR),
            ("Generated At", data.generated_at),
            ("Scan Name", data.scan_name),
            ("Scan ID", data.scan_id),
            ("History ID", data.history_id),
            ("Source File", data.source_file),
        ]
        row = 3
        for k, v in meta_rows:
            ws.cell(row, 1, k).font = Font(bold=True)
            ws.cell(row, 2, v)
            row += 1

        metric_start = 3
        ws.cell(metric_start, 4, "Metric").font = Font(bold=True)
        ws.cell(metric_start, 5, "Value").font = Font(bold=True)
        ws.cell(metric_start, 4).fill = fills["header"]
        ws.cell(metric_start, 5).fill = fills["header"]
        for idx, (k, v) in enumerate(data.metrics.items(), start=metric_start + 1):
            ws.cell(idx, 4, k)
            ws.cell(idx, 5, v)
            ws.cell(idx, 4).border = border
            ws.cell(idx, 5).border = border

        # Chart source table for key status only.
        chart_row = 18
        ws.cell(chart_row, 1, "Status").font = Font(bold=True)
        ws.cell(chart_row, 2, "Count").font = Font(bold=True)
        status_metric_map = [
            ("Auth Passed", data.metrics.get("Auth Passed", 0)),
            ("Auth Failed", data.metrics.get("Auth Failed", 0)),
            ("Partial Auth", data.metrics.get("Partial Auth", 0)),
            ("No Credentials", data.metrics.get("No Credentials", 0)),
            ("Not Reachable", data.metrics.get("Not Reachable", 0)),
            ("Unknown", data.metrics.get("Unknown", 0)),
        ]
        for i, (label, count) in enumerate(status_metric_map, start=chart_row + 1):
            ws.cell(i, 1, label)
            ws.cell(i, 2, count)

        pie = PieChart()
        labels = Reference(ws, min_col=1, min_row=chart_row + 1, max_row=chart_row + len(status_metric_map))
        values = Reference(ws, min_col=2, min_row=chart_row, max_row=chart_row + len(status_metric_map))
        pie.add_data(values, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Authentication Status"
        ws.add_chart(pie, "D18")

        # Failure reason chart table.
        reason_counts = Counter()
        for h in data.host_records:
            if h.status in {AuthStatus.FAIL, AuthStatus.PARTIAL, AuthStatus.NOCREDS, AuthStatus.NOT_REACHABLE, AuthStatus.UNKNOWN}:
                if h.reasons:
                    reason_counts[h.reasons[0]] += 1
        reason_start = 18
        ws.cell(reason_start, 8, "Failure / Issue Reason").font = Font(bold=True)
        ws.cell(reason_start, 9, "Count").font = Font(bold=True)
        for i, (reason, count) in enumerate(reason_counts.most_common(10), start=reason_start + 1):
            ws.cell(i, 8, reason[:90])
            ws.cell(i, 9, count)
        if reason_counts:
            bar = BarChart()
            data_ref = Reference(ws, min_col=9, min_row=reason_start, max_row=reason_start + min(10, len(reason_counts)))
            cats_ref = Reference(ws, min_col=8, min_row=reason_start + 1, max_row=reason_start + min(10, len(reason_counts)))
            bar.add_data(data_ref, titles_from_data=True)
            bar.set_categories(cats_ref)
            bar.title = "Top Authentication Issues"
            bar.height = 7
            bar.width = 14
            ws.add_chart(bar, "D35")

        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 24

        def add_table_sheet(title: str, rows: List[Dict[str, Any]]) -> None:
            ws2 = wb.create_sheet(title[:31])
            if not rows:
                ws2.append(["No data"])
                return
            headers = list(rows[0].keys())
            ws2.append(headers)
            for c in range(1, len(headers) + 1):
                cell = ws2.cell(1, c)
                cell.font = Font(bold=True)
                cell.fill = fills["header"]
                cell.border = border
            for row_data in rows:
                ws2.append([row_data.get(h, "") for h in headers])
            ws2.freeze_panes = "A2"
            ws2.auto_filter.ref = ws2.dimensions
            for c, header in enumerate(headers, start=1):
                max_len = max([len(str(header))] + [len(str(r.get(header, ""))) for r in rows[:200]])
                ws2.column_dimensions[get_column_letter(c)].width = min(max(max_len + 2, 12), 60)
            # Conditional status fills.
            if "Status" in headers:
                s_col = headers.index("Status") + 1
                for r in range(2, ws2.max_row + 1):
                    val = str(ws2.cell(r, s_col).value or "")
                    fill = None
                    if val == "PASS": fill = fills["pass"]
                    elif val == "FAIL": fill = fills["fail"]
                    elif val == "PARTIAL": fill = fills["partial"]
                    elif val == "NOCREDS": fill = fills["nocreds"]
                    elif val == "NOT_REACHABLE": fill = fills["not_reachable"]
                    elif val == "UNKNOWN": fill = fills["unknown"]
                    if fill:
                        for c in range(1, ws2.max_column + 1):
                            ws2.cell(r, c).fill = fill

        host_rows = Exporter.host_rows(data)
        protocol_rows = Exporter.protocol_rows(data)
        finding_rows = Exporter.finding_rows(data)
        add_table_sheet("Host_Status", host_rows)
        add_table_sheet("Protocol_Status", protocol_rows)
        add_table_sheet("Auth_Findings", finding_rows)
        for status in STATUS_ORDER:
            add_table_sheet(f"{status.value}_Hosts", [r for r in host_rows if r["Status"] == status.value])

        notes_ws = wb.create_sheet("Notes")
        notes_ws.append(["Notes"])
        notes_ws["A1"].font = Font(bold=True)
        for note in data.notes:
            notes_ws.append([note])
        notes_ws.column_dimensions["A"].width = 100

        wb.save(path)

    @staticmethod
    def export_pdf(data: DashboardData, path: Path) -> None:
        if not MATPLOTLIB_AVAILABLE:
            raise RuntimeError("matplotlib not installed. Install: python3 -m pip install matplotlib")

        with PdfPages(str(path)) as pdf:
            fig = Figure(figsize=(11.69, 8.27))
            ax = fig.add_subplot(111)
            ax.axis("off")
            title = f"{APP_NAME}\nScan: {data.scan_name} | Generated: {data.generated_at}"
            lines = [title, "", "Summary Metrics"]
            for k, v in data.metrics.items():
                lines.append(f"  {k}: {v}")
            lines.extend(["", "Notes"])
            lines.extend([f"  - {n}" for n in data.notes] or ["  - None"])
            ax.text(0.02, 0.98, "\n".join(lines), va="top", ha="left", fontsize=11, family="monospace")
            pdf.savefig(fig)

            # Pie chart page.
            fig2 = Figure(figsize=(11.69, 8.27))
            ax2 = fig2.add_subplot(111)
            labels = ["Passed", "Failed", "Partial", "No Creds", "Not Reachable", "Unknown"]
            sizes = [
                data.metrics.get("Auth Passed", 0),
                data.metrics.get("Auth Failed", 0),
                data.metrics.get("Partial Auth", 0),
                data.metrics.get("No Credentials", 0),
                data.metrics.get("Not Reachable", 0),
                data.metrics.get("Unknown", 0),
            ]
            if sum(sizes) > 0:
                ax2.pie(sizes, labels=labels, autopct="%1.1f%%")
            ax2.set_title("Authentication Status Breakdown")
            pdf.savefig(fig2)

            # Top issue reasons.
            reason_counts = Counter()
            for h in data.host_records:
                if h.status != AuthStatus.PASS:
                    reason_counts[(h.reasons[0] if h.reasons else "Unknown")[:80]] += 1
            fig3 = Figure(figsize=(11.69, 8.27))
            ax3 = fig3.add_subplot(111)
            if reason_counts:
                items = reason_counts.most_common(10)
                labels3 = [x[0] for x in items][::-1]
                counts3 = [x[1] for x in items][::-1]
                ax3.barh(labels3, counts3)
                ax3.set_xlabel("Host Count")
            ax3.set_title("Top Authentication Issues")
            fig3.tight_layout()
            pdf.savefig(fig3)

            # Failed / Partial / NoCreds table page.
            fig4 = Figure(figsize=(11.69, 8.27))
            ax4 = fig4.add_subplot(111)
            ax4.axis("off")
            rows = [h for h in data.host_records if h.status != AuthStatus.PASS][:35]
            table_lines = ["Host Status Details (first 35 non-pass hosts)", ""]
            for h in rows:
                table_lines.append(f"{h.host:20} {h.status.value:9} Plugins: {','.join(map(str,h.plugin_ids))[:30]} Reason: {' | '.join(h.reasons)[:80]}")
            if not rows:
                table_lines.append("All hosts are passed or no non-pass evidence available.")
            ax4.text(0.02, 0.98, "\n".join(table_lines), va="top", ha="left", fontsize=9, family="monospace")
            pdf.savefig(fig4)

# -----------------------------
# Local login
# -----------------------------

class LocalAuthManager:
    def __init__(self, path: Path = AUTH_CONFIG_PATH):
        self.path = path

    def is_configured(self) -> bool:
        return self.path.exists()

    def load(self) -> Dict[str, Any]:
        try:
            return json.loads(self.path.read_text(encoding="utf-8"))
        except Exception:
            return {}

    def configure(self, username: str, password: str) -> None:
        username = username.strip()
        if not username:
            raise ValueError("Username is required.")
        if len(password) < 8:
            raise ValueError("Password must be at least 8 characters.")
        salt = secrets.token_bytes(16)
        password_hash = self.hash_password(password, salt)
        payload = {
            "username": username,
            "salt": base64.b64encode(salt).decode("ascii"),
            "password_hash": base64.b64encode(password_hash).decode("ascii"),
            "iterations": AUTH_ITERATIONS,
            "created_at": now_string(),
        }
        self.path.write_text(json.dumps(payload, indent=2), encoding="utf-8")

    def hash_password(self, password: str, salt: bytes, iterations: int = AUTH_ITERATIONS) -> bytes:
        return hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, iterations)

    def verify(self, username: str, password: str) -> bool:
        payload = self.load()
        if username.strip() != str(payload.get("username", "")):
            return False
        try:
            salt = base64.b64decode(payload.get("salt", ""))
            expected = base64.b64decode(payload.get("password_hash", ""))
            iterations = int(payload.get("iterations", AUTH_ITERATIONS))
        except Exception:
            return False
        actual = self.hash_password(password, salt, iterations)
        return hmac.compare_digest(actual, expected)


class LoginDialog:
    def __init__(self, root: tk.Tk, auth: LocalAuthManager):
        self.root = root
        self.auth = auth
        self.authenticated = False
        self.setup_mode = not auth.is_configured()
        self.username_var = tk.StringVar()
        self.password_var = tk.StringVar()
        self.confirm_var = tk.StringVar()
        self.logo_phase = 0
        self.logo_after_id: Optional[str] = None

        self.window = tk.Toplevel(root)
        self.window.title(f"{APP_NAME} Login")
        self.window.resizable(False, False)
        self.window.configure(bg="#F8FAFC")
        self.window.transient(root)
        self.window.grab_set()
        self.window.protocol("WM_DELETE_WINDOW", self.cancel)

        self.build()
        self.center()
        self.username_entry.focus_set()

    def build(self):
        self.window.geometry("980x560")
        shell = tk.Frame(self.window, bg="#F8FAFC")
        shell.pack(fill="both", expand=True)

        left = tk.Frame(shell, bg="#F8FAFC", width=420)
        left.pack(side="left", fill="both")
        left.pack_propagate(False)

        form = tk.Frame(left, bg="#F8FAFC")
        form.place(relx=0.5, rely=0.5, anchor="center", width=310)

        tk.Label(form, text="TRINETRA", bg="#F8FAFC", fg="#0F766E", font=("Segoe UI", 20, "bold")).pack(anchor="w")
        tk.Label(
            form,
            text="Trinetra Credential Assurance",
            bg="#F8FAFC",
            fg="#334155",
            font=("Segoe UI", 10, "bold"),
        ).pack(anchor="w", pady=(2, 26))

        title = "Create secure access" if self.setup_mode else "Welcome back"
        subtitle = "Set up your local admin login" if self.setup_mode else "Sign in to continue remediation work"
        tk.Label(form, text=title, bg="#F8FAFC", fg="#0F172A", font=("Segoe UI", 19, "bold")).pack(anchor="w")
        tk.Label(form, text=subtitle, bg="#F8FAFC", fg="#64748B", font=("Segoe UI", 10)).pack(anchor="w", pady=(4, 22))

        tk.Label(form, text="Username", bg="#F8FAFC", fg="#334155", font=("Segoe UI", 9, "bold")).pack(anchor="w")
        self.username_entry = tk.Entry(form, textvariable=self.username_var, width=34, bg="#FFFFFF", fg="#0F172A", relief="solid", bd=1, font=("Segoe UI", 11))
        self.username_entry.pack(fill="x", ipady=8, pady=(4, 14))

        tk.Label(form, text="Password", bg="#F8FAFC", fg="#334155", font=("Segoe UI", 9, "bold")).pack(anchor="w")
        tk.Entry(form, textvariable=self.password_var, width=34, show="*", bg="#FFFFFF", fg="#0F172A", relief="solid", bd=1, font=("Segoe UI", 11)).pack(fill="x", ipady=8, pady=(4, 14))

        if self.setup_mode:
            tk.Label(form, text="Confirm password", bg="#F8FAFC", fg="#334155", font=("Segoe UI", 9, "bold")).pack(anchor="w")
            tk.Entry(form, textvariable=self.confirm_var, width=34, show="*", bg="#FFFFFF", fg="#0F172A", relief="solid", bd=1, font=("Segoe UI", 11)).pack(fill="x", ipady=8, pady=(4, 14))

        button_text = "Create Login" if self.setup_mode else "Sign In"
        tk.Button(
            form,
            text=button_text,
            command=self.submit,
            bg="#0F766E",
            fg="#FFFFFF",
            activebackground="#115E59",
            activeforeground="#FFFFFF",
            relief="flat",
            bd=0,
            font=("Segoe UI", 11, "bold"),
            cursor="hand2",
        ).pack(fill="x", ipady=9, pady=(4, 10))

        tk.Button(
            form,
            text="Exit",
            command=self.cancel,
            bg="#E2E8F0",
            fg="#0F172A",
            activebackground="#CBD5E1",
            activeforeground="#0F172A",
            relief="flat",
            bd=0,
            font=("Segoe UI", 10, "bold"),
            cursor="hand2",
        ).pack(fill="x", ipady=8)

        self.message_var = tk.StringVar()
        tk.Label(form, textvariable=self.message_var, bg="#F8FAFC", fg="#DC2626", font=("Segoe UI", 9)).pack(anchor="w", pady=(10, 0))

        tk.Label(
            form,
            text="Local authentication protects dashboard access on this workstation.",
            bg="#F8FAFC",
            fg="#64748B",
            wraplength=300,
            justify="left",
            font=("Segoe UI", 9),
        ).pack(anchor="w", pady=(24, 0))

        right = tk.Canvas(shell, width=560, height=560, bg="#111827", highlightthickness=0)
        right.pack(side="right", fill="both", expand=True)
        self.hero_canvas = right
        self.draw_login_art()

        self.window.bind("<Return>", lambda _e: self.submit())

    def draw_login_art(self):
        c = self.hero_canvas
        for x in range(0, 580, 28):
            c.create_line(x, 0, x, 560, fill="#1F2937")
        for y in range(0, 580, 28):
            c.create_line(0, y, 560, y, fill="#1F2937")
        c.create_oval(350, 350, 760, 760, fill="#0F766E", outline="")
        c.create_oval(390, 390, 690, 690, fill="#1D4ED8", outline="")
        c.create_text(68, 78, text="Trinetra", fill="#5EEAD4", font=("Segoe UI", 13, "bold"), anchor="w")
        c.create_text(
            68,
            160,
            text="CONTROL\nWHAT SCANNERS\nCAN VERIFY",
            fill="#F8FAFC",
            font=("Segoe UI", 34, "bold"),
            anchor="w",
            justify="left",
        )
        c.create_text(
            70,
            330,
            text="Turn authentication evidence into focused remediation.\nPrioritize failed, partial, missing-credential, and\nunreachable targets before they become audit gaps.",
            fill="#CBD5E1",
            font=("Segoe UI", 15, "bold"),
            anchor="w",
            justify="left",
            width=430,
        )
        c.create_text(70, 470, text="\"Visibility becomes value when it drives ownership.\"", fill="#A7F3D0", font=("Segoe UI", 13, "italic"), anchor="w")
        c.create_text(70, 505, text=f"Built by {AUTHOR}", fill="#94A3B8", font=("Segoe UI", 10, "bold"), anchor="w")

        self.logo_ring = c.create_oval(398, 70, 492, 164, outline="#5EEAD4", width=3)
        self.logo_ring_inner = c.create_oval(416, 88, 474, 146, outline="#38BDF8", width=2)
        self.logo_dot = c.create_oval(440, 64, 454, 78, fill="#F8FAFC", outline="")
        self.logo_core = c.create_text(445, 118, text="CA", fill="#F8FAFC", font=("Segoe UI", 17, "bold"))
        self.animate_logo()

    def animate_logo(self):
        if not hasattr(self, "hero_canvas"):
            return
        self.logo_phase += 1
        angle = self.logo_phase / 12
        radius = 50
        cx, cy = 445, 117
        x = cx + math.cos(angle) * radius
        y = cy + math.sin(angle) * radius
        self.hero_canvas.coords(self.logo_dot, x - 7, y - 7, x + 7, y + 7)
        pulse = 4 + math.sin(angle * 1.5) * 4
        self.hero_canvas.coords(self.logo_ring_inner, 416 - pulse, 88 - pulse, 474 + pulse, 146 + pulse)
        self.logo_after_id = self.window.after(50, self.animate_logo)

    def center(self):
        self.window.update_idletasks()
        width = self.window.winfo_width()
        height = self.window.winfo_height()
        x = self.root.winfo_screenwidth() // 2 - width // 2
        y = self.root.winfo_screenheight() // 2 - height // 2
        self.window.geometry(f"+{x}+{y}")

    def submit(self):
        username = self.username_var.get().strip()
        password = self.password_var.get()
        try:
            if self.setup_mode:
                if password != self.confirm_var.get():
                    raise ValueError("Passwords do not match.")
                self.auth.configure(username, password)
                self.authenticated = True
            else:
                if not self.auth.verify(username, password):
                    raise ValueError("Invalid username or password.")
                self.authenticated = True
            if self.logo_after_id:
                try:
                    self.window.after_cancel(self.logo_after_id)
                except Exception:
                    pass
            self.window.destroy()
        except Exception as exc:
            self.message_var.set(str(exc))

    def cancel(self):
        if self.logo_after_id:
            try:
                self.window.after_cancel(self.logo_after_id)
            except Exception:
                pass
        self.authenticated = False
        self.window.destroy()


# -----------------------------
# GUI Application
# -----------------------------

class NessusAuthDashboardGUI:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title(f"{APP_NAME} v{APP_VERSION}")
        self.root.geometry("1360x850")
        self.root.minsize(1150, 700)

        self.data: Optional[DashboardData] = None
        self.scans: List[Dict[str, Any]] = []
        self.scan_folders: List[Dict[str, Any]] = []
        self.scan_histories: List[Dict[str, Any]] = []
        self.history_label_to_id: Dict[str, str] = {}
        self.dark_mode = tk.BooleanVar(value=True)
        self.verify_tls = tk.BooleanVar(value=False)
        self.base_url_var = tk.StringVar(value="https://127.0.0.1:8834")
        self.access_key_var = tk.StringVar()
        self.secret_key_var = tk.StringVar()
        self.history_id_var = tk.StringVar()
        self.history_filter_var = tk.StringVar()
        self.filter_text_var = tk.StringVar()
        self.filter_status_var = tk.StringVar(value="ALL")
        self.selected_scan_count_var = tk.StringVar(value="Selected scans: 0")
        self.status_var = tk.StringVar(value="Ready")
        self.progress_var = tk.IntVar(value=0)

        self.selected_folder_id: str = ""
        self.selected_scan_id: Optional[str] = None
        self.selected_scan_name: str = ""
        self.selected_scans: List[Tuple[str, str]] = []
        self.last_output_folder: Optional[Path] = None

        self.style = ttk.Style()
        self.setup_style()
        self.build_ui()
        self.apply_theme()

    # ---------- UI setup ----------

    def setup_style(self):
        try:
            self.style.theme_use("clam")
        except Exception:
            pass

    def apply_theme(self):
        dark = self.dark_mode.get()
        if dark:
            bg = "#101828"
            fg = "#F8FAFC"
            panel = "#182235"
            entry = "#243246"
            header = "#0B1220"
            accent = "#06B6D4"
            muted = "#A5B4FC"
            border = "#334155"
            text_bg = "#0F172A"
            text_fg = "#E2E8F0"
            tab_selected = "#0E7490"
            button_bg = "#2563EB"
            button_active = "#1D4ED8"
            card_palette = {
                "Total": "#0E7490",
                "Pass": STATUS_COLORS[AuthStatus.PASS],
                "Fail": STATUS_COLORS[AuthStatus.FAIL],
                "Partial": STATUS_COLORS[AuthStatus.PARTIAL],
                "NoCreds": STATUS_COLORS[AuthStatus.NOCREDS],
                "NotReachable": STATUS_COLORS[AuthStatus.NOT_REACHABLE],
                "Unknown": STATUS_COLORS[AuthStatus.UNKNOWN],
                "Coverage": "#14B8A6",
                "Success": "#84CC16",
            }
        else:
            bg = "#F8FAFC"
            fg = "#0F172A"
            panel = "#FFFFFF"
            entry = "#FFFFFF"
            header = "#DBEAFE"
            accent = "#0891B2"
            muted = "#4F46E5"
            border = "#CBD5E1"
            text_bg = "#FFFFFF"
            text_fg = "#0F172A"
            tab_selected = "#38BDF8"
            button_bg = "#2563EB"
            button_active = "#1D4ED8"
            card_palette = {
                "Total": "#06B6D4",
                "Pass": "#16A34A",
                "Fail": "#DC2626",
                "Partial": "#EA580C",
                "NoCreds": "#7C3AED",
                "NotReachable": "#475569",
                "Unknown": "#CA8A04",
                "Coverage": "#0D9488",
                "Success": "#65A30D",
            }

        self.root.configure(bg=bg)
        self.style.configure("TFrame", background=bg)
        self.style.configure("Panel.TFrame", background=panel)
        self.style.configure("TLabel", background=bg, foreground=fg)
        self.style.configure("Panel.TLabel", background=panel, foreground=fg)
        self.style.configure("AppTitle.TLabel", background=bg, foreground=accent, font=("Segoe UI", 18, "bold"))
        self.style.configure("Subtitle.TLabel", background=bg, foreground=muted, font=("Segoe UI", 10))
        self.style.configure("Version.TLabel", background=panel, foreground=fg, font=("Segoe UI", 9, "bold"), padding=[8, 3])
        self.style.configure("Header.TLabel", background=bg, foreground=accent, font=("Segoe UI", 15, "bold"))
        self.style.configure("Card.TFrame", background=panel, relief="ridge", borderwidth=1)
        self.style.configure("CardTitle.TLabel", background=panel, foreground=muted, font=("Segoe UI", 10, "bold"))
        self.style.configure("CardValue.TLabel", background=panel, foreground=fg, font=("Segoe UI", 24, "bold"))
        self.style.configure("TButton", padding=6, background=button_bg, foreground="#FFFFFF", bordercolor=button_bg)
        self.style.map("TButton", background=[("active", button_active), ("pressed", button_active)], foreground=[("disabled", "#94A3B8")])
        self.style.configure("TCheckbutton", background=bg, foreground=fg)
        self.style.configure("TEntry", fieldbackground=entry, foreground=fg, bordercolor=border)
        self.style.configure("TCombobox", fieldbackground=entry, foreground=fg, background=entry, arrowcolor=accent)
        self.style.configure("TLabelFrame", background=bg, foreground=accent, bordercolor=border)
        self.style.configure("TLabelFrame.Label", background=bg, foreground=accent, font=("Segoe UI", 9, "bold"))
        self.style.configure("Treeview", background=entry, foreground=fg, fieldbackground=entry, rowheight=26)
        self.style.configure("Treeview.Heading", background=header, foreground=fg, font=("Segoe UI", 9, "bold"), bordercolor=accent)
        self.style.configure("TNotebook", background=bg)
        self.style.configure("TNotebook.Tab", padding=[10, 5], background=panel, foreground=fg)
        self.style.map("TNotebook.Tab", background=[("selected", tab_selected)], foreground=[("selected", "#FFFFFF")])
        self.style.configure("Horizontal.TProgressbar", troughcolor=panel, background=accent, bordercolor=border, lightcolor=accent, darkcolor=accent)

        card_names = ["Total", "Pass", "Fail", "Partial", "NoCreds", "NotReachable", "Unknown", "Coverage", "Success"]
        for name in card_names:
            color = card_palette[name]
            self.style.configure(f"{name}Card.TFrame", background=color, relief="flat")
            self.style.configure(f"{name}CardTitle.TLabel", background=color, foreground="#FFFFFF", font=("Segoe UI", 10, "bold"))
            self.style.configure(f"{name}CardValue.TLabel", background=color, foreground="#FFFFFF", font=("Segoe UI", 24, "bold"))

        for widget_name in ("summary_text", "drill_text", "log_text"):
            widget = getattr(self, widget_name, None)
            if widget is not None:
                widget.configure(bg=text_bg, fg=text_fg, insertbackground=accent, relief="flat", padx=10, pady=8)

        for tree_name in ("folder_tree", "scan_tree", "host_tree", "protocol_tree", "finding_tree"):
            tree = getattr(self, tree_name, None)
            if tree is not None:
                self.configure_tree_tags(tree)

        if getattr(self, "data", None):
            self.draw_charts()

    def build_ui(self):
        self.build_top_bar()
        self.build_scan_section()
        self.build_action_bar()
        self.build_dashboard_tabs()
        self.build_status_bar()

    def build_top_bar(self):
        frame = ttk.Frame(self.root)
        frame.pack(fill="x", padx=12, pady=(10, 6))

        title_group = ttk.Frame(frame)
        title_group.pack(side="left", fill="x", expand=True)
        ttk.Label(title_group, text=APP_NAME, style="AppTitle.TLabel").pack(anchor="w")
        ttk.Label(
            title_group,
            text=f"Credentialed scan coverage, authentication status, and reachability reporting | Author: {AUTHOR}",
            style="Subtitle.TLabel",
        ).pack(anchor="w", pady=(2, 0))

        ttk.Label(frame, text=f"v{APP_VERSION}", style="Version.TLabel").pack(side="right", padx=(10, 0))
        ttk.Checkbutton(frame, text="Dark Mode", variable=self.dark_mode, command=self.apply_theme).pack(side="right")

    def build_scan_section(self):
        container = ttk.LabelFrame(self.root, text="Connection and Scan Selection")
        container.pack(fill="x", padx=10, pady=4)

        connection_row = ttk.Frame(container)
        connection_row.pack(fill="x", padx=8, pady=(4, 2))
        ttk.Label(connection_row, text="Base URL").pack(side="left")
        ttk.Entry(connection_row, textvariable=self.base_url_var, width=34).pack(side="left", padx=5)
        ttk.Checkbutton(connection_row, text="Verify TLS", variable=self.verify_tls).pack(side="left", padx=6)
        ttk.Button(connection_row, text="Test / Load Folders", command=self.load_scans_thread).pack(side="left", padx=4)
        ttk.Button(connection_row, text="Offline: Load CSV", command=self.load_offline_csv).pack(side="left", padx=4)

        credential_row = ttk.Frame(container)
        credential_row.pack(fill="x", padx=8, pady=(2, 4))
        ttk.Label(credential_row, text="Access Key").pack(side="left")
        ttk.Entry(credential_row, textvariable=self.access_key_var, width=48, show="*").pack(side="left", padx=5)
        ttk.Label(credential_row, text="Secret Key").pack(side="left")
        ttk.Entry(credential_row, textvariable=self.secret_key_var, width=48, show="*").pack(side="left", padx=5)

        row2 = ttk.Frame(container)
        row2.pack(fill="both", expand=True, padx=8, pady=4)

        folder_panel = ttk.Frame(row2)
        folder_panel.pack(side="left", fill="y", padx=(0, 8))
        ttk.Label(folder_panel, text="1. Folder", style="Panel.TLabel").pack(anchor="w", pady=(0, 3))
        self.folder_tree = ttk.Treeview(folder_panel, columns=("name", "count", "folder_id"), displaycolumns=("name", "count"), show="headings", height=6)
        self.folder_tree.heading("name", text="Folder")
        self.folder_tree.heading("count", text="Scans")
        self.folder_tree.column("name", width=240, anchor="w")
        self.folder_tree.column("count", width=55, anchor="center")
        self.folder_tree.pack(side="left", fill="y")
        self.folder_tree.bind("<<TreeviewSelect>>", self.on_folder_select)
        folder_scroll = ttk.Scrollbar(folder_panel, orient="vertical", command=self.folder_tree.yview)
        self.folder_tree.configure(yscrollcommand=folder_scroll.set)
        folder_scroll.pack(side="right", fill="y")

        scan_panel = ttk.Frame(row2)
        scan_panel.pack(side="left", fill="both", expand=True)
        ttk.Label(scan_panel, text="2. Scan (select one or more)", style="Panel.TLabel").pack(anchor="w", pady=(0, 3))
        columns = ("id", "name", "status", "folder", "last_modification_date")
        self.scan_tree = ttk.Treeview(scan_panel, columns=columns, show="headings", height=6, selectmode="extended")
        for col, label, width in [
            ("id", "Scan ID", 80),
            ("name", "Scan Name", 520),
            ("status", "Status", 110),
            ("folder", "Folder", 180),
            ("last_modification_date", "Last Modified", 160),
        ]:
            self.scan_tree.heading(col, text=label)
            self.scan_tree.column(col, width=width, anchor="w")
        self.scan_tree.pack(side="left", fill="both", expand=True)
        self.scan_tree.bind("<<TreeviewSelect>>", self.on_scan_select)
        scroll = ttk.Scrollbar(scan_panel, orient="vertical", command=self.scan_tree.yview)
        self.scan_tree.configure(yscrollcommand=scroll.set)
        scroll.pack(side="right", fill="y")

        row3 = ttk.Frame(container)
        row3.pack(fill="x", padx=8, pady=4)
        ttk.Button(row3, text="Select All Visible Scans", command=self.select_all_visible_scans).pack(side="left", padx=(0, 6))
        ttk.Button(row3, text="Clear Scan Selection", command=self.clear_scan_selection).pack(side="left", padx=(0, 12))
        ttk.Label(row3, textvariable=self.selected_scan_count_var, style="Panel.TLabel").pack(side="left", padx=(0, 14))
        ttk.Label(row3, text="3. Scan History").pack(side="left")
        self.history_combo = ttk.Combobox(row3, textvariable=self.history_filter_var, width=58, state="readonly", values=[])
        self.history_combo.pack(side="left", padx=6)
        self.history_combo.bind("<<ComboboxSelected>>", self.on_history_select)
        ttk.Button(row3, text="Build Dashboard", command=self.load_preview_thread).pack(side="left", padx=6)
        ttk.Button(row3, text="Clear", command=self.clear_dashboard).pack(side="left", padx=4)

    def build_action_bar(self):
        frame = ttk.Frame(self.root)
        frame.pack(fill="x", padx=10, pady=4)
        filter_group = ttk.Frame(frame)
        filter_group.pack(side="left", fill="x", expand=True)
        ttk.Label(filter_group, text="Search").pack(side="left")
        entry = ttk.Entry(filter_group, textvariable=self.filter_text_var, width=35)
        entry.pack(side="left", padx=5)
        entry.bind("<KeyRelease>", lambda _e: self.refresh_host_table())
        ttk.Label(filter_group, text="Status").pack(side="left")
        status_combo = ttk.Combobox(filter_group, textvariable=self.filter_status_var, width=14, state="readonly",
                                    values=["ALL"] + [s.value for s in STATUS_ORDER])
        status_combo.pack(side="left", padx=5)
        status_combo.bind("<<ComboboxSelected>>", lambda _e: self.refresh_host_table())

        export_group = ttk.Frame(frame)
        export_group.pack(side="right")
        ttk.Button(export_group, text="Export Excel", command=self.export_excel).pack(side="left", padx=3)
        ttk.Button(export_group, text="Export PDF", command=self.export_pdf).pack(side="left", padx=3)
        ttk.Button(export_group, text="Export CSV Bundle", command=self.export_csv_bundle).pack(side="left", padx=3)
        ttk.Button(export_group, text="Copy Failed IPs", command=lambda: self.copy_ips(AuthStatus.FAIL)).pack(side="left", padx=3)
        ttk.Button(export_group, text="Open Output Folder", command=self.open_output_folder).pack(side="left", padx=3)

    def build_dashboard_tabs(self):
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=4)

        self.dashboard_tab = ttk.Frame(self.notebook)
        self.hosts_tab = ttk.Frame(self.notebook)
        self.protocols_tab = ttk.Frame(self.notebook)
        self.findings_tab = ttk.Frame(self.notebook)
        self.drilldown_tab = ttk.Frame(self.notebook)
        self.logs_tab = ttk.Frame(self.notebook)

        self.notebook.add(self.dashboard_tab, text="Dashboard")
        self.notebook.add(self.hosts_tab, text="Host Status")
        self.notebook.add(self.protocols_tab, text="Protocol Breakdown")
        self.notebook.add(self.findings_tab, text="Auth Findings")
        self.notebook.add(self.drilldown_tab, text="Host Drilldown")
        self.notebook.add(self.logs_tab, text="Logs")

        self.build_dashboard_tab()
        self.build_hosts_tab()
        self.build_protocols_tab()
        self.build_findings_tab()
        self.build_drilldown_tab()
        self.build_logs_tab()

    def build_dashboard_tab(self):
        top = ttk.Frame(self.dashboard_tab)
        top.pack(fill="x", padx=8, pady=8)
        for col in range(5):
            top.columnconfigure(col, weight=1, uniform="metric_cards")
        self.card_labels: Dict[str, tk.Label] = {}
        card_defs = [
            ("Total IPs", "Total IPs"),
            ("Auth Passed", "Auth Passed"),
            ("Auth Failed", "Auth Failed"),
            ("Partial Auth", "Partial Auth"),
            ("No Credentials", "No Credentials"),
            ("Not Reachable", "Not Reachable"),
            ("Unknown", "Unknown"),
            ("Credential Coverage %", "Coverage %"),
            ("Auth Success %", "Success %"),
        ]
        for idx, (metric_key, title) in enumerate(card_defs):
            frame_style, title_style, value_style = METRIC_CARD_STYLES.get(
                metric_key,
                ("Card.TFrame", "CardTitle.TLabel", "CardValue.TLabel"),
            )
            card = ttk.Frame(top, style=frame_style)
            card.grid(row=idx // 5, column=idx % 5, sticky="ew", padx=4, pady=4)
            ttk.Label(card, text=title, style=title_style).pack(anchor="w", padx=10, pady=(8, 0))
            lbl = ttk.Label(card, text="0", style=value_style)
            lbl.pack(anchor="w", padx=10, pady=(0, 8))
            self.card_labels[metric_key] = lbl

        middle = ttk.Frame(self.dashboard_tab)
        middle.pack(fill="both", expand=True, padx=8, pady=8)
        self.chart_frame_left = ttk.LabelFrame(middle, text="Authentication Status Breakdown")
        self.chart_frame_left.pack(side="left", fill="both", expand=True, padx=4)
        self.chart_frame_right = ttk.LabelFrame(middle, text="Top Authentication Issues and Protocols")
        self.chart_frame_right.pack(side="left", fill="both", expand=True, padx=4)

        self.chart_canvas_left = None
        self.chart_canvas_right = None

        self.summary_text = tk.Text(self.dashboard_tab, height=6, wrap="word")
        self.summary_text.pack(fill="x", padx=8, pady=8)
        self.summary_text.insert("end", "Dashboard will appear here after loading preview.\n")
        self.summary_text.config(state="disabled")

    def build_tree_with_scroll(self, parent, columns: Tuple[str, ...], headings: Dict[str, str], widths: Dict[str, int]) -> ttk.Treeview:
        frame = ttk.Frame(parent)
        frame.pack(fill="both", expand=True, padx=8, pady=8)
        tree = ttk.Treeview(frame, columns=columns, show="headings")
        for c in columns:
            tree.heading(c, text=headings.get(c, c))
            tree.column(c, width=widths.get(c, 120), anchor="w")
        vs = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        hs = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vs.set, xscrollcommand=hs.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vs.grid(row=0, column=1, sticky="ns")
        hs.grid(row=1, column=0, sticky="ew")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)
        return tree

    def build_hosts_tab(self):
        columns = ("Host", "Source Scans", "Status", "Pass", "Fail", "Partial", "NoCreds", "Unknown", "Plugin IDs", "Reason", "Recommendation")
        headings = {c: c for c in columns}
        widths = {
            "Host": 150, "Source Scans": 220, "Status": 90, "Pass": 120, "Fail": 120, "Partial": 120, "NoCreds": 120,
            "Unknown": 120, "Plugin IDs": 150, "Reason": 300, "Recommendation": 420
        }
        self.host_tree = self.build_tree_with_scroll(self.hosts_tab, columns, headings, widths)
        self.host_tree.bind("<<TreeviewSelect>>", self.on_host_select)

    def build_protocols_tab(self):
        columns = ("Host", "Source Scans", "Auth Protocol", "Status", "Plugin IDs", "Reasons", "Accounts", "Ports", "Confidence", "Evidence Count")
        headings = {c: c for c in columns}
        widths = {"Host": 150, "Source Scans": 220, "Auth Protocol": 120, "Status": 90, "Plugin IDs": 150, "Reasons": 360, "Accounts": 180, "Ports": 120, "Confidence": 180, "Evidence Count": 110}
        self.protocol_tree = self.build_tree_with_scroll(self.protocols_tab, columns, headings, widths)

    def build_findings_tab(self):
        columns = ("Host", "Source Scan", "Auth Protocol", "Plugin ID", "Plugin Name", "Evidence Type", "Reason", "Port", "Account")
        headings = {c: c for c in columns}
        widths = {"Host": 150, "Source Scan": 220, "Auth Protocol": 120, "Plugin ID": 90, "Plugin Name": 350, "Evidence Type": 160, "Reason": 330, "Port": 80, "Account": 180}
        self.finding_tree = self.build_tree_with_scroll(self.findings_tab, columns, headings, widths)
        self.finding_tree.bind("<<TreeviewSelect>>", self.on_finding_select)

    def build_drilldown_tab(self):
        top = ttk.Frame(self.drilldown_tab)
        top.pack(fill="x", padx=8, pady=8)
        ttk.Label(top, text="Selected Host / Finding Evidence", style="Header.TLabel").pack(side="left")
        self.drill_text = tk.Text(self.drilldown_tab, wrap="word")
        self.drill_text.pack(fill="both", expand=True, padx=8, pady=8)
        self.drill_text.insert("end", "Select a host or finding to view detailed evidence.\n")
        self.drill_text.config(state="disabled")

    def build_logs_tab(self):
        self.log_text = tk.Text(self.logs_tab, wrap="word")
        self.log_text.pack(fill="both", expand=True, padx=8, pady=8)
        self.log("Application started.")

    def build_status_bar(self):
        frame = ttk.Frame(self.root)
        frame.pack(fill="x", padx=10, pady=(0, 8))
        ttk.Label(frame, textvariable=self.status_var).pack(side="left")
        self.progress = ttk.Progressbar(frame, variable=self.progress_var, maximum=100, length=260)
        self.progress.pack(side="right")

    # ---------- Logging/status ----------

    def log(self, msg: str):
        timestamp = now_string()
        try:
            self.log_text.insert("end", f"[{timestamp}] {msg}\n")
            self.log_text.see("end")
        except Exception:
            pass

    def thread_log(self, msg: str):
        self.root.after(0, lambda: self.log(msg))
        self.root.after(0, lambda: self.status_var.set(msg))

    def set_progress(self, value: int):
        self.root.after(0, lambda: self.progress_var.set(max(0, min(100, value))))

    def show_error(self, title: str, message: str):
        self.root.after(0, lambda: messagebox.showerror(title, message))
        self.thread_log(f"ERROR: {message}")

    # ---------- Actions ----------

    def make_client(self) -> NessusClient:
        base_url = self.base_url_var.get().strip()
        access_key = self.access_key_var.get().strip()
        secret_key = self.secret_key_var.get().strip()
        if not base_url:
            raise ValueError("Base URL is required.")
        if not access_key or not secret_key:
            raise ValueError("Access Key and Secret Key are required.")
        return NessusClient(base_url, access_key, secret_key, verify_tls=self.verify_tls.get())

    def load_scans_thread(self):
        threading.Thread(target=self._load_scans_worker, daemon=True).start()

    def _load_scans_worker(self):
        try:
            self.thread_log("Connecting to Nessus and loading folders...")
            self.set_progress(10)
            client = self.make_client()
            scans, folders = client.list_scan_inventory()
            self.set_progress(70)
            self.scans = scans
            self.scan_folders = folders
            self.root.after(0, self.populate_folder_selector)
            self.thread_log(f"Loaded {len(folders)} folders and {len(scans)} scans.")
            self.set_progress(100)
        except Exception as exc:
            self.show_error("Load Scans Failed", str(exc))
            self.thread_log(traceback.format_exc())
            self.set_progress(0)

    def populate_folder_selector(self):
        for item in self.folder_tree.get_children():
            self.folder_tree.delete(item)

        first_item = ""
        for folder in self.scan_folders:
            folder_id = str(folder.get("id", ""))
            folder_name = str(folder.get("name", "") or f"Folder {folder_id}")
            count = sum(1 for scan in self.scans if str(scan.get("folder_id", "")) == folder_id)
            item = self.folder_tree.insert("", "end", values=(folder_name, count, folder_id))
            if not first_item:
                first_item = item

        known_folder_ids = {str(folder.get("id", "")) for folder in self.scan_folders}
        missing_folder_count = sum(1 for scan in self.scans if str(scan.get("folder_id", "")) not in known_folder_ids)
        if missing_folder_count:
            item = self.folder_tree.insert("", "end", values=("Unfiled / Unknown Folder", missing_folder_count, "__MISSING__"))
            if not first_item:
                first_item = item

        if not first_item and self.scans:
            first_item = self.folder_tree.insert("", "end", values=("All Scans", len(self.scans), ""))

        if first_item:
            self.folder_tree.selection_set(first_item)
            self.folder_tree.focus(first_item)
            self.on_folder_select()
        else:
            self.selected_folder_id = ""
            self.populate_scan_tree()

    def on_folder_select(self, _event=None):
        sel = self.folder_tree.selection()
        if sel:
            values = self.folder_tree.item(sel[0], "values")
            self.selected_folder_id = str(values[2]) if len(values) >= 3 else ""
        self.populate_scan_tree()

    def update_selected_scan_count(self):
        count = len(self.selected_scans)
        self.selected_scan_count_var.set(f"Selected scans: {count}")

    def populate_scan_tree(self):
        for item in self.scan_tree.get_children():
            self.scan_tree.delete(item)
        self.selected_scan_id = None
        self.selected_scan_name = ""
        self.selected_scans = []
        self.update_selected_scan_count()
        self.clear_history_selector()
        folder_id = self.selected_folder_id
        if folder_id == "__MISSING__":
            known_folder_ids = {str(folder.get("id", "")) for folder in self.scan_folders}
            visible_scans = [s for s in self.scans if str(s.get("folder_id", "")) not in known_folder_ids]
        else:
            visible_scans = [
                s for s in self.scans
                if not folder_id or str(s.get("folder_id", "")) == folder_id
            ]
        for s in visible_scans:
            scan_id = str(s.get("id", ""))
            name = str(s.get("name", ""))
            status = str(s.get("status", ""))
            folder = str(s.get("folder_name", s.get("folder_id", "")))
            lm = format_nessus_time(s.get("last_modification_date", ""))
            self.scan_tree.insert("", "end", values=(scan_id, name, status, folder, lm))
        folder_name = "selected folder"
        folder_sel = self.folder_tree.selection()
        if folder_sel:
            values = self.folder_tree.item(folder_sel[0], "values")
            if values:
                folder_name = str(values[0])
        self.status_var.set(f"Showing {len(visible_scans)} scans in {folder_name}")

    def on_scan_select(self, _event=None):
        sel = self.scan_tree.selection()
        if not sel:
            self.selected_scan_id = None
            self.selected_scan_name = ""
            self.selected_scans = []
            self.update_selected_scan_count()
            self.clear_history_selector()
            return

        selected_scans: List[Tuple[str, str]] = []
        for item in sel:
            values = self.scan_tree.item(item, "values")
            if len(values) >= 2 and values[0]:
                selected_scans.append((str(values[0]), str(values[1])))

        self.selected_scans = selected_scans
        self.update_selected_scan_count()
        if not selected_scans:
            self.selected_scan_id = None
            self.selected_scan_name = ""
            self.status_var.set("No valid scan selected")
            return

        if len(selected_scans) > 1:
            self.selected_scan_id = None
            self.selected_scan_name = ""
            self.clear_history_selector()
            multi_history_label = "Latest/default history for each selected scan"
            self.history_combo.configure(values=[multi_history_label])
            self.history_filter_var.set(multi_history_label)
            self.status_var.set(f"Selected {len(selected_scans)} scans. Build Dashboard will merge latest/default results.")
            return

        self.selected_scan_id, self.selected_scan_name = selected_scans[0]
        self.status_var.set(f"Selected scan {self.selected_scan_id}: {self.selected_scan_name}. Loading histories...")
        self.clear_history_selector()
        threading.Thread(target=self._load_histories_worker, args=(self.selected_scan_id, self.selected_scan_name), daemon=True).start()

    def select_all_visible_scans(self):
        items = self.scan_tree.get_children()
        if not items:
            return
        self.scan_tree.selection_set(items)
        self.on_scan_select()

    def clear_scan_selection(self):
        self.scan_tree.selection_remove(self.scan_tree.selection())
        self.selected_scan_id = None
        self.selected_scan_name = ""
        self.selected_scans = []
        self.update_selected_scan_count()
        self.clear_history_selector()
        self.status_var.set("Scan selection cleared.")

    def clear_history_selector(self):
        self.scan_histories = []
        self.history_label_to_id = {}
        self.history_id_var.set("")
        self.history_filter_var.set("")
        if hasattr(self, "history_combo"):
            self.history_combo.configure(values=[])

    def _load_histories_worker(self, scan_id: str, scan_name: str):
        try:
            client = self.make_client()
            details = client.get_scan_details(scan_id)
            histories = details.get("history", []) or details.get("histories", []) or []
            self.root.after(0, lambda: self.populate_history_selector(histories, scan_id, scan_name))
        except Exception as exc:
            self.root.after(0, lambda: self.populate_history_selector([], scan_id, scan_name))
            self.thread_log(f"Could not load scan histories for {scan_id}: {exc}")

    def history_label(self, history: Dict[str, Any]) -> Tuple[str, str]:
        history_id = str(history.get("history_id") or history.get("id") or "")
        status = str(history.get("status") or history.get("readable_status") or "history")
        timestamp = history.get("last_modification_date") or history.get("creation_date") or history.get("created_at") or ""
        timestamp = format_nessus_time(timestamp)
        label = f"{timestamp or 'Saved history'} | {status}"
        if history_id:
            label = f"{label} | ID {history_id}"
        return label, history_id

    def populate_history_selector(self, histories: List[Dict[str, Any]], scan_id: str, scan_name: str):
        if self.selected_scan_id != scan_id:
            return
        self.scan_histories = histories
        self.history_label_to_id = {}
        labels = []
        for history in histories:
            label, history_id = self.history_label(history)
            if label in self.history_label_to_id:
                label = f"{label} #{len(labels) + 1}"
            labels.append(label)
            self.history_label_to_id[label] = history_id

        if not labels:
            labels = ["Latest/default history"]
            self.history_label_to_id[labels[0]] = ""

        self.history_combo.configure(values=labels)
        self.history_filter_var.set(labels[0])
        self.on_history_select()
        self.status_var.set(f"Selected scan {scan_id}: {scan_name}. Choose history, then build dashboard.")

    def on_history_select(self, _event=None):
        label = self.history_filter_var.get()
        self.history_id_var.set(self.history_label_to_id.get(label, ""))

    def load_preview_thread(self):
        if not self.selected_scans and self.selected_scan_id:
            self.selected_scans = [(self.selected_scan_id, self.selected_scan_name or self.selected_scan_id)]
        if not self.selected_scans:
            messagebox.showwarning("No Scan Selected", "Please select one or more scans first or use Offline CSV.")
            return
        threading.Thread(target=self._load_preview_worker, daemon=True).start()

    def _load_preview_worker(self):
        try:
            selected_scans = list(self.selected_scans)
            single_scan = len(selected_scans) == 1
            history_id = self.history_id_var.get().strip() if single_scan else ""
            classifier = AuthClassifier()

            if single_scan:
                self.thread_log(f"Loading dashboard preview for scan {selected_scans[0][0]}...")
            else:
                self.thread_log(f"Loading combined dashboard preview for {len(selected_scans)} scans...")
            self.set_progress(5)
            client = self.make_client()

            authoritative_hosts: List[str] = []
            configured_targets: List[str] = []
            all_findings: List[AuthFinding] = []
            raw_rows_count = 0
            scan_names: List[str] = []
            scan_ids: List[str] = []

            with tempfile.TemporaryDirectory(prefix="nessus_auth_preview_") as tmpdir:
                total_scans = max(len(selected_scans), 1)
                for index, (scan_id, scan_name) in enumerate(selected_scans, start=1):
                    scan_names.append(scan_name or scan_id)
                    scan_ids.append(scan_id)
                    current_history_id = history_id if single_scan else ""
                    self.thread_log(f"[{index}/{total_scans}] Fetching scan details for {scan_id}: {scan_name or scan_id}")
                    try:
                        details = client.get_scan_details(scan_id, current_history_id or None)
                        detail_hosts = classifier.hosts_from_scan_details(details)
                        detail_targets = classifier.targets_from_scan_details(details)
                        authoritative_hosts.extend(detail_hosts)
                        configured_targets.extend(detail_targets)
                        classifier.remember_host_sources(detail_hosts, scan_id, scan_name or scan_id)
                        classifier.remember_target_sources(detail_targets, scan_id, scan_name or scan_id)
                        self.thread_log(f"[{index}/{total_scans}] Host inventory loaded.")
                    except Exception as exc:
                        self.thread_log(f"[{index}/{total_scans}] Could not fetch scan details; will fallback to CSV hosts. Reason: {exc}")

                    self.set_progress(10 + int(index / total_scans * 20))
                    csv_path = Path(tmpdir) / f"nessus_scan_{scan_id}_preview.csv"
                    client.export_scan_csv(scan_id, csv_path, current_history_id or None, progress_cb=self.thread_log)
                    raw_rows, findings, csv_hosts = classifier.parse_csv(csv_path, scan_id=scan_id, scan_name=scan_name or scan_id)
                    authoritative_hosts.extend(csv_hosts)
                    classifier.remember_host_sources(csv_hosts, scan_id, scan_name or scan_id)
                    all_findings.extend(findings)
                    raw_rows_count += len(raw_rows)
                    self.thread_log(f"[{index}/{total_scans}] Parsed {len(raw_rows)} raw rows and {len(findings)} auth-related findings.")
                    self.set_progress(30 + int(index / total_scans * 45))

                classifier.configured_targets = sort_hosts(configured_targets)
                self.set_progress(80)
                dashboard_scan_name = scan_names[0] if single_scan else f"Combined scans ({len(selected_scans)})"
                dashboard_scan_id = scan_ids[0] if single_scan else ",".join(scan_ids)
                dashboard_history_id = history_id if single_scan else "latest/default per scan"
                source_file = "Temporary Nessus API CSV export" if single_scan else "Temporary Nessus API CSV exports"
                data = classifier.classify(
                    scan_name=dashboard_scan_name,
                    scan_id=dashboard_scan_id,
                    history_id=dashboard_history_id,
                    authoritative_hosts=sort_hosts(authoritative_hosts),
                    findings=all_findings,
                    raw_rows_count=raw_rows_count,
                    source_file=source_file,
                )

            self.data = data
            self.root.after(0, self.render_dashboard)
            self.thread_log("Dashboard preview ready. You can now export Excel/PDF/CSV.")
            self.set_progress(100)
        except Exception as exc:
            self.show_error("Preview Failed", str(exc))
            self.thread_log(traceback.format_exc())
            self.set_progress(0)

    def load_offline_csv(self):
        path = filedialog.askopenfilename(
            title="Select Nessus CSV Export",
            filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")]
        )
        if not path:
            return
        try:
            classifier = AuthClassifier()
            raw_rows, findings, csv_hosts = classifier.parse_csv(Path(path), scan_id="offline", scan_name=Path(path).stem)
            classifier.remember_host_sources(csv_hosts, "offline", Path(path).stem)
            data = classifier.classify(
                scan_name=Path(path).stem,
                scan_id="offline",
                history_id="",
                authoritative_hosts=csv_hosts,
                findings=findings,
                raw_rows_count=len(raw_rows),
                source_file=path,
            )
            self.data = data
            self.render_dashboard()
            self.log(f"Offline CSV loaded: {path}")
            self.status_var.set("Offline dashboard preview ready.")
            self.progress_var.set(100)
        except Exception as exc:
            messagebox.showerror("Offline CSV Failed", str(exc))
            self.log(traceback.format_exc())

    def clear_dashboard(self):
        self.data = None
        self.progress_var.set(0)
        self.status_var.set("Cleared")
        for tree in [self.host_tree, self.protocol_tree, self.finding_tree]:
            for item in tree.get_children():
                tree.delete(item)
        for lbl in self.card_labels.values():
            lbl.config(text="0")
        self.set_text(self.summary_text, "Dashboard cleared.\n")
        self.set_text(self.drill_text, "Select a host or finding to view detailed evidence.\n")
        self.clear_charts()

    # ---------- Rendering ----------

    def render_dashboard(self):
        if not self.data:
            return
        data = self.data
        for key, lbl in self.card_labels.items():
            value = data.metrics.get(key, 0)
            if isinstance(value, float):
                lbl.config(text=f"{value}%" if "%" in key else str(value))
            else:
                lbl.config(text=str(value))
        self.refresh_host_table()
        self.refresh_protocol_table()
        self.refresh_finding_table()
        self.refresh_summary_text()
        self.draw_charts()
        self.notebook.select(self.dashboard_tab)

    def refresh_summary_text(self):
        if not self.data:
            return
        m = self.data.metrics
        lines = [
            f"Scan Name: {self.data.scan_name}",
            f"Scan ID: {self.data.scan_id} | History ID: {self.data.history_id or 'latest/default'} | Generated: {self.data.generated_at}",
            "",
            f"Mutually exclusive counts: Total={m.get('Total IPs',0)}, Pass={m.get('Auth Passed',0)}, Fail={m.get('Auth Failed',0)}, Partial={m.get('Partial Auth',0)}, No Credentials={m.get('No Credentials',0)}, Not Reachable={m.get('Not Reachable',0)}, Unknown={m.get('Unknown',0)}",
            f"Credential Coverage % = Pass + Partial / Total = {m.get('Credential Coverage %',0)}%",
            f"Auth Success % = Pass / Total = {m.get('Auth Success %',0)}%",
            "",
            "Important: Unknown is kept as a diagnostic state to avoid false Pass/Fail counting.",
        ]
        if self.data.notes:
            lines.append("Notes:")
            lines += [f"- {n}" for n in self.data.notes]
        self.set_text(self.summary_text, "\n".join(lines))

    def set_text(self, widget: tk.Text, text: str):
        widget.config(state="normal")
        widget.delete("1.0", "end")
        widget.insert("end", text)
        widget.config(state="disabled")

    def refresh_host_table(self):
        if not self.data:
            return
        for item in self.host_tree.get_children():
            self.host_tree.delete(item)
        filter_text = self.filter_text_var.get().strip().lower()
        filter_status = self.filter_status_var.get().strip()
        for h in self.data.host_records:
            if filter_status != "ALL" and h.status.value != filter_status:
                continue
            searchable = " ".join([h.host, h.status.value, " ".join(h.scan_names), " ".join(h.scan_ids), " ".join(h.reasons), " ".join(map(str, h.plugin_ids))]).lower()
            if filter_text and filter_text not in searchable:
                continue
            vals = (
                h.host,
                ", ".join(h.scan_names or h.scan_ids),
                h.status.value,
                ", ".join(h.pass_protocols),
                ", ".join(h.fail_protocols),
                ", ".join(h.partial_protocols),
                ", ".join(h.nocreds_protocols),
                ", ".join(h.unknown_protocols),
                ", ".join(map(str, h.plugin_ids)),
                " | ".join(h.reasons)[:500],
                h.recommendation,
            )
            self.host_tree.insert("", "end", values=vals, tags=(h.status.value,))
        self.configure_tree_tags(self.host_tree)

    def refresh_protocol_table(self):
        if not self.data:
            return
        for item in self.protocol_tree.get_children():
            self.protocol_tree.delete(item)
        for p in self.data.protocol_records:
            vals = (
                p.host,
                ", ".join(p.scan_names or p.scan_ids),
                p.auth_protocol,
                p.status.value,
                ", ".join(map(str, p.plugin_ids)),
                " | ".join(p.reasons)[:500],
                ", ".join(p.accounts),
                ", ".join(p.ports),
                p.confidence,
                p.evidence_count,
            )
            self.protocol_tree.insert("", "end", values=vals, tags=(p.status.value,))
        self.configure_tree_tags(self.protocol_tree)

    def refresh_finding_table(self):
        if not self.data:
            return
        for item in self.finding_tree.get_children():
            self.finding_tree.delete(item)
        for f in self.data.findings:
            vals = (
                f.host,
                f.scan_name or f.scan_id or f.source,
                f.auth_protocol,
                f.plugin_id,
                f.plugin_name,
                f.evidence_type,
                f.reason,
                f.port if f.port is not None else "",
                f.account,
            )
            status_tag = self.finding_status_tag(f)
            self.finding_tree.insert("", "end", values=vals, tags=(status_tag,))
        self.configure_tree_tags(self.finding_tree)

    def finding_status_tag(self, f: AuthFinding) -> str:
        if f.plugin_id in PRIMARY_PASS or (f.plugin_id in SECONDARY_PASS and (f.plugin_id != 19506 or credentialed_checks_yes(f.plugin_output))):
            return "PASS"
        if f.plugin_id in PRIMARY_FAIL:
            return "FAIL"
        if f.plugin_id in PRIMARY_PARTIAL:
            return "PARTIAL"
        if f.plugin_id in PRIMARY_NOCREDS:
            return "NOCREDS"
        return "UNKNOWN"

    def configure_tree_tags(self, tree: ttk.Treeview):
        try:
            if self.dark_mode.get():
                tree.tag_configure("PASS", background="#14532D", foreground="#DCFCE7")
                tree.tag_configure("FAIL", background="#7F1D1D", foreground="#FEE2E2")
                tree.tag_configure("PARTIAL", background="#7C2D12", foreground="#FFEDD5")
                tree.tag_configure("NOCREDS", background="#4C1D95", foreground="#EDE9FE")
                tree.tag_configure("NOT_REACHABLE", background="#334155", foreground="#F8FAFC")
                tree.tag_configure("UNKNOWN", background="#713F12", foreground="#FEF9C3")
            else:
                tree.tag_configure("PASS", background="#BBF7D0", foreground="#052E16")
                tree.tag_configure("FAIL", background="#FECACA", foreground="#450A0A")
                tree.tag_configure("PARTIAL", background="#FED7AA", foreground="#431407")
                tree.tag_configure("NOCREDS", background="#DDD6FE", foreground="#2E1065")
                tree.tag_configure("NOT_REACHABLE", background="#CBD5E1", foreground="#0F172A")
                tree.tag_configure("UNKNOWN", background="#FEF08A", foreground="#422006")
        except Exception:
            pass

    def clear_charts(self):
        for canvas in [self.chart_canvas_left, self.chart_canvas_right]:
            if canvas:
                try:
                    canvas.get_tk_widget().destroy()
                except Exception:
                    pass
        self.chart_canvas_left = None
        self.chart_canvas_right = None

    def draw_charts(self):
        if not self.data or not MATPLOTLIB_AVAILABLE:
            if not MATPLOTLIB_AVAILABLE:
                self.thread_log("matplotlib not installed; charts disabled. Install: python3 -m pip install matplotlib")
            return
        self.clear_charts()
        data = self.data

        dark = self.dark_mode.get()
        figure_bg = "#182235" if dark else "#FFFFFF"
        axis_fg = "#E2E8F0" if dark else "#0F172A"

        fig1 = Figure(figsize=(5.7, 3.8), dpi=100, facecolor=figure_bg)
        ax1 = fig1.add_subplot(111)
        ax1.set_facecolor(figure_bg)
        labels = ["Pass", "Fail", "Partial", "No Creds", "Not Reachable", "Unknown"]
        sizes = [
            data.metrics.get("Auth Passed", 0),
            data.metrics.get("Auth Failed", 0),
            data.metrics.get("Partial Auth", 0),
            data.metrics.get("No Credentials", 0),
            data.metrics.get("Not Reachable", 0),
            data.metrics.get("Unknown", 0),
        ]
        if sum(sizes) > 0:
            pie_colors = [
                STATUS_COLORS[AuthStatus.PASS],
                STATUS_COLORS[AuthStatus.FAIL],
                STATUS_COLORS[AuthStatus.PARTIAL],
                STATUS_COLORS[AuthStatus.NOCREDS],
                STATUS_COLORS[AuthStatus.NOT_REACHABLE],
                STATUS_COLORS[AuthStatus.UNKNOWN],
            ]
            ax1.pie(sizes, labels=labels, autopct="%1.1f%%", startangle=90, colors=pie_colors)
        ax1.set_title("Authentication Status")
        ax1.title.set_color(axis_fg)
        fig1.tight_layout()
        self.chart_canvas_left = FigureCanvasTkAgg(fig1, master=self.chart_frame_left)
        self.chart_canvas_left.draw()
        self.chart_canvas_left.get_tk_widget().pack(fill="both", expand=True)

        fig2 = Figure(figsize=(5.7, 3.8), dpi=100, facecolor=figure_bg)
        ax2 = fig2.add_subplot(111)
        ax2.set_facecolor(figure_bg)
        reason_counts = Counter()
        protocol_issue_counts = Counter()
        for h in data.host_records:
            if h.status != AuthStatus.PASS:
                reason_counts[(h.reasons[0] if h.reasons else "Unknown")[:45]] += 1
        for p in data.protocol_records:
            if p.status != AuthStatus.PASS:
                protocol_issue_counts[p.auth_protocol] += 1
        if reason_counts:
            items = reason_counts.most_common(8)
            names = [x[0] for x in items][::-1]
            counts = [x[1] for x in items][::-1]
            ax2.barh(names, counts, color="#EF4444")
            ax2.set_xlabel("Host Count")
            ax2.set_title("Top Auth Issues")
        else:
            items = protocol_issue_counts.most_common(8)
            names = [x[0] for x in items][::-1]
            counts = [x[1] for x in items][::-1]
            ax2.barh(names, counts, color="#06B6D4")
            ax2.set_xlabel("Count")
            ax2.set_title("Protocol Issues")
        ax2.title.set_color(axis_fg)
        ax2.xaxis.label.set_color(axis_fg)
        ax2.tick_params(axis="x", colors=axis_fg)
        ax2.tick_params(axis="y", colors=axis_fg)
        for spine in ax2.spines.values():
            spine.set_color(axis_fg)
        fig2.tight_layout()
        self.chart_canvas_right = FigureCanvasTkAgg(fig2, master=self.chart_frame_right)
        self.chart_canvas_right.draw()
        self.chart_canvas_right.get_tk_widget().pack(fill="both", expand=True)

    # ---------- Drilldown ----------

    def on_host_select(self, _event=None):
        if not self.data:
            return
        sel = self.host_tree.selection()
        if not sel:
            return
        values = self.host_tree.item(sel[0], "values")
        if not values:
            return
        host = str(values[0])
        hrec = next((h for h in self.data.host_records if h.host == host), None)
        if not hrec:
            return
        related_protocols = [p for p in self.data.protocol_records if p.host == host]
        related_findings = [f for f in self.data.findings if f.host == host]
        lines = [
            f"Host: {hrec.host}",
            f"Source Scans: {', '.join(hrec.scan_names or hrec.scan_ids) or 'N/A'}",
            f"Final Status: {hrec.status.value}",
            f"Recommendation: {hrec.recommendation}",
            "",
            "Protocol Status:",
        ]
        for p in related_protocols:
            source = ", ".join(p.scan_names or p.scan_ids) or "N/A"
            lines.append(f"  - {p.auth_protocol}: {p.status.value} | Source: {source} | Plugins: {', '.join(map(str,p.plugin_ids))} | Reasons: {' | '.join(p.reasons)}")
        if not related_protocols:
            lines.append("  - No protocol-level authentication evidence found.")
        lines.extend(["", "Raw Evidence:"])
        for f in related_findings[:25]:
            lines.append("=" * 90)
            lines.append(f"Source Scan: {f.scan_name or 'N/A'} ({f.scan_id or 'no scan id'})")
            lines.append(f"Plugin {f.plugin_id} - {f.plugin_name}")
            lines.append(f"Protocol: {f.auth_protocol} | Port: {f.port or ''} | Account: {f.account or ''}")
            lines.append(f"Reason: {f.reason}")
            if f.plugin_output:
                lines.append("Plugin Output:")
                lines.append(f.plugin_output[:2000])
        if len(related_findings) > 25:
            lines.append(f"\n... {len(related_findings)-25} more findings available in Auth Findings tab.")
        self.set_text(self.drill_text, "\n".join(lines))
        self.notebook.select(self.drilldown_tab)

    def on_finding_select(self, _event=None):
        if not self.data:
            return
        sel = self.finding_tree.selection()
        if not sel:
            return
        values = self.finding_tree.item(sel[0], "values")
        if len(values) < 4:
            return
        host = str(values[0])
        source_scan = str(values[1])
        plugin_id = safe_int(values[3])
        if plugin_id is None:
            return
        candidates = [
            f for f in self.data.findings
            if f.host == host and f.plugin_id == plugin_id and source_scan in {f.scan_name, f.scan_id, f.source}
        ]
        if not candidates:
            candidates = [f for f in self.data.findings if f.host == host and f.plugin_id == plugin_id]
        if not candidates:
            return
        f = candidates[0]
        lines = [
            f"Host: {f.host}",
            f"Source Scan: {f.scan_name or 'N/A'}",
            f"Source Scan ID: {f.scan_id or 'N/A'}",
            f"Plugin ID: {f.plugin_id}",
            f"Plugin Name: {f.plugin_name}",
            f"Auth Protocol: {f.auth_protocol}",
            f"Port: {f.port or ''}",
            f"Account: {f.account or ''}",
            f"Evidence Type: {f.evidence_type}",
            f"Reason: {f.reason}",
            "",
            "Plugin Output:",
            f.plugin_output or "No plugin output available in CSV.",
        ]
        self.set_text(self.drill_text, "\n".join(lines))
        self.notebook.select(self.drilldown_tab)

    # ---------- Exports ----------

    def require_data(self) -> DashboardData:
        if not self.data:
            raise RuntimeError("Load dashboard preview first.")
        return self.data

    def export_excel(self):
        try:
            data = self.require_data()
            default = f"nessus_auth_dashboard_{data.scan_id}_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            path = filedialog.asksaveasfilename(title="Save Excel Dashboard", defaultextension=".xlsx", initialfile=default,
                                                filetypes=[("Excel Workbook", "*.xlsx")])
            if not path:
                return
            Exporter.export_excel(data, Path(path))
            self.last_output_folder = Path(path).parent
            messagebox.showinfo("Export Complete", f"Excel dashboard saved:\n{path}")
            self.log(f"Excel exported: {path}")
        except Exception as exc:
            messagebox.showerror("Excel Export Failed", str(exc))
            self.log(traceback.format_exc())

    def export_pdf(self):
        try:
            data = self.require_data()
            default = f"nessus_auth_dashboard_{data.scan_id}_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            path = filedialog.asksaveasfilename(title="Save PDF Dashboard", defaultextension=".pdf", initialfile=default,
                                                filetypes=[("PDF", "*.pdf")])
            if not path:
                return
            Exporter.export_pdf(data, Path(path))
            self.last_output_folder = Path(path).parent
            messagebox.showinfo("Export Complete", f"PDF dashboard saved:\n{path}")
            self.log(f"PDF exported: {path}")
        except Exception as exc:
            messagebox.showerror("PDF Export Failed", str(exc))
            self.log(traceback.format_exc())

    def export_csv_bundle(self):
        try:
            data = self.require_data()
            folder = filedialog.askdirectory(title="Select Folder for CSV Bundle")
            if not folder:
                return
            out = Path(folder) / f"nessus_auth_csv_bundle_{data.scan_id}_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}"
            Exporter.export_csv_bundle(data, out)
            self.last_output_folder = out
            messagebox.showinfo("Export Complete", f"CSV bundle saved:\n{out}")
            self.log(f"CSV bundle exported: {out}")
        except Exception as exc:
            messagebox.showerror("CSV Export Failed", str(exc))
            self.log(traceback.format_exc())

    def copy_ips(self, status: AuthStatus):
        try:
            data = self.require_data()
            ips = [h.host for h in data.host_records if h.status == status]
            self.root.clipboard_clear()
            self.root.clipboard_append("\n".join(ips))
            self.root.update()
            messagebox.showinfo("Copied", f"Copied {len(ips)} {status.value} IPs/hosts to clipboard.")
        except Exception as exc:
            messagebox.showerror("Copy Failed", str(exc))

    def open_output_folder(self):
        folder = self.last_output_folder
        if not folder or not folder.exists():
            messagebox.showinfo("No Folder", "No output folder available yet. Export first.")
            return
        try:
            if os.name == "nt":
                os.startfile(str(folder))  # type: ignore[attr-defined]
            elif sys.platform == "darwin":
                subprocess.Popen(["open", str(folder)])
            else:
                subprocess.Popen(
                    ["xdg-open", str(folder)],
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL,
                )
        except Exception as exc:
            messagebox.showerror("Open Folder Failed", str(exc))

# -----------------------------
# Entry point
# -----------------------------

def main():
    root = tk.Tk()
    root.withdraw()
    auth = LocalAuthManager()
    login = LoginDialog(root, auth)
    root.wait_window(login.window)
    if not login.authenticated:
        root.destroy()
        return
    root.deiconify()
    app = NessusAuthDashboardGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
